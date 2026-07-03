"""
Automation Ambassador Program – Centralized Tool Monitor
Manages: Dashboard Factory (8080), Dashboard Factory QA (5555), Jira Tracker (8082)
Run via:  Launch Monitor.command   OR   ./venv/bin/uvicorn monitor:app --port 9000
"""

import asyncio
import base64
import calendar
import io
import json
import os
import platform
import signal
import subprocess
import sys
import tempfile
import threading
import time
import uuid
from datetime import datetime, date, timezone
from pathlib import Path
from typing import Dict, Optional

import httpx
import psutil
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, StreamingResponse
from pydantic import BaseModel

_IS_WIN  = platform.system() == "Windows"
_SCRIPTS = "Scripts" if _IS_WIN else "bin"
_PY      = "python"  if _IS_WIN else "python3"
_UVI     = "uvicorn"

# ---------------------------------------------------------------------------
# Tool registry
# ---------------------------------------------------------------------------

BASE = Path(__file__).parent

# Make export_calculated_fields importable from its original location
_CF_DIR = str(BASE / "01_Dashboard_Factory")
if _CF_DIR not in sys.path:
    sys.path.insert(0, _CF_DIR)

CF_OUTPUT_DIR = BASE / "calc_fields_output"
CF_OUTPUT_DIR.mkdir(exist_ok=True)
CF_JOBS: Dict[str, dict] = {}  # job_id -> {status, path, count, filename, error}

CONFIG_FILE = BASE / "monitor_config.json"

def _load_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text())
        except Exception:
            pass
    return {}

def _save_config(data: dict) -> None:
    existing = _load_config()
    existing.update(data)
    CONFIG_FILE.write_text(json.dumps(existing, indent=2))

TOOLS = {
    "dashboard-factory": {
        "name": "Dashboard Factory",
        "description": "Tableau workbook generation from Postgres",
        "port": 8080,
        "url": "http://localhost:8080",
        "cwd": str(BASE / "01_Dashboard_Factory"),
        "cmd": [f"venv/{_SCRIPTS}/{_PY}", "-m", "uvicorn", "app:app", "--host", "127.0.0.1", "--port", "8080"],
        "health_path": "/docs",
        "color": "#4F9CF9",
        "icon": "🏭",
    },
    "dashboard-factory-qa": {
        "name": "Dashboard Factory QA",
        "description": "Tableau workbook formatting & QA compliance checks",
        "port": 5555,
        "url": "http://localhost:5555",
        "cwd": str(BASE / "02_Dashboard_Factory_QA" / "Tableau QA Compliance"),
        "cmd": [str(BASE / "02_Dashboard_Factory_QA" / ".venv" / _SCRIPTS / _PY), "app.py"],
        "health_path": "/",
        "color": "#A78BFA",
        "icon": "✅",
    },
    "jira-tracker": {
        "name": "Jira Tracker",
        "description": "Jira activity tracking & timesheet export",
        "port": 8082,
        "url": "http://localhost:8082",
        "cwd": str(BASE / "03_Jira_Tracker"),
        "cmd": [f"venv/{_SCRIPTS}/{_PY}", "-m", "uvicorn", "jira_tracker:app", "--host", "127.0.0.1", "--port", "8082"],
        "health_path": "/docs",
        "color": "#34D399",
        "icon": "📋",
    },
}

# Process registry: tool_id -> {proc, started_at, log_lines}
PROCS: Dict[str, dict] = {}
LOG_BUFFER = 500

# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------

app = FastAPI(title="Automation Ambassador Program Monitor", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _capture_output(tool_id: str, proc: subprocess.Popen) -> None:
    """Background threads to drain stdout/stderr into the log buffer."""
    def _read(stream):
        for raw in iter(stream.readline, b""):
            line = raw.decode("utf-8", errors="replace").rstrip()
            entry = PROCS.get(tool_id)
            if entry is None:
                break
            buf: list = entry["log_lines"]
            buf.append(line)
            if len(buf) > LOG_BUFFER:
                buf.pop(0)

    if proc.stdout:
        threading.Thread(target=_read, args=(proc.stdout,), daemon=True).start()
    if proc.stderr:
        threading.Thread(target=_read, args=(proc.stderr,), daemon=True).start()


async def _health(tool_id: str) -> bool:
    tool = TOOLS[tool_id]
    try:
        async with httpx.AsyncClient(timeout=2.0) as client:
            r = await client.get(f"http://localhost:{tool['port']}{tool['health_path']}")
            return r.status_code < 500
    except Exception:
        return False


def _is_managed_alive(tool_id: str) -> bool:
    entry = PROCS.get(tool_id)
    return entry is not None and entry["proc"].poll() is None


def _psutil_kill(pid: int, killed: list) -> None:
    """Kill a process and all its descendants using psutil (cross-platform)."""
    try:
        proc = psutil.Process(pid)
        for child in proc.children(recursive=True):
            try:
                child.kill()
                killed.append(child.pid)
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
        proc.kill()
        killed.append(pid)
    except (psutil.NoSuchProcess, psutil.AccessDenied):
        pass


async def _kill_by_port(port: int) -> list:
    """Kill all processes listening on port and their children."""
    listening_pids: set = set()
    if _IS_WIN:
        try:
            for c in psutil.net_connections(kind="inet"):
                if c.laddr.port == port and c.status == psutil.CONN_LISTEN and c.pid:
                    listening_pids.add(c.pid)
        except (psutil.AccessDenied, Exception):
            pass
    else:
        try:
            out = subprocess.check_output(
                ["lsof", "-t", "-i", f"TCP:{port}", "-s", "TCP:LISTEN"],
                text=True, stderr=subprocess.DEVNULL,
            )
            listening_pids = {int(p) for p in out.split() if p.strip().isdigit()}
        except subprocess.CalledProcessError:
            pass

    if not listening_pids:
        return []

    pid_set = set(listening_pids)
    roots, leaves = [], []
    for pid in listening_pids:
        try:
            ppid = psutil.Process(pid).ppid()
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            ppid = 0
        (leaves if ppid in pid_set else roots).append(pid)

    killed: list = []
    for pid in roots + leaves:
        _psutil_kill(pid, killed)
    return killed


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

_VERSION_FILE = BASE / "VERSION"

def _local_version() -> str:
    try:
        return _VERSION_FILE.read_text().strip()
    except Exception:
        return "unknown"

@app.get("/", response_class=HTMLResponse)
async def root():
    html = Path(__file__).parent / "monitor.html"
    return HTMLResponse(html.read_text(encoding="utf-8"))


@app.get("/api/version")
async def api_version():
    local = _local_version()
    latest = None
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            r = await client.get(
                "https://api.github.com/repos/Rajkumar9750/automation-ambassador-program/releases/latest",
                headers={"Accept": "application/vnd.github+json"}
            )
            if r.status_code == 200:
                latest = r.json().get("tag_name", "").lstrip("v")
    except Exception:
        pass
    update_available = bool(latest and latest != local)
    return {"version": local, "latest": latest, "update_available": update_available}


@app.get("/api/tools")
async def api_tools():
    checks = await asyncio.gather(*[_health(tid) for tid in TOOLS])
    results = []
    for (tool_id, tool), is_up in zip(TOOLS.items(), checks):
        managed = _is_managed_alive(tool_id)
        entry = PROCS.get(tool_id)
        uptime = int(time.time() - entry["started_at"]) if managed and entry else None
        if is_up:
            status = "running"
        elif managed:
            status = "starting"
        else:
            status = "stopped"
        results.append({
            "id": tool_id,
            "name": tool["name"],
            "description": tool["description"],
            "port": tool["port"],
            "url": tool["url"],
            "color": tool["color"],
            "icon": tool["icon"],
            "status": status,
            "managed": managed,
            "uptime": uptime,
            "pid": entry["proc"].pid if managed and entry else None,
            "started_at": entry["started_at"] if managed and entry else None,
        })
    return results


@app.post("/api/tools/{tool_id}/start")
async def api_start(tool_id: str):
    if tool_id not in TOOLS:
        return JSONResponse({"error": "Unknown tool"}, status_code=404)
    if _is_managed_alive(tool_id):
        return {"status": "already_running", "pid": PROCS[tool_id]["proc"].pid}

    tool = TOOLS[tool_id]
    popen_kwargs: dict = dict(cwd=tool["cwd"], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if _IS_WIN:
        popen_kwargs["creationflags"] = subprocess.CREATE_NEW_PROCESS_GROUP
    else:
        popen_kwargs["preexec_fn"] = os.setsid
    try:
        proc = subprocess.Popen(tool["cmd"], **popen_kwargs)
    except FileNotFoundError as e:
        return JSONResponse({"error": f"Could not launch {tool['name']}: {e}. Try re-running install.ps1."}, status_code=500)
    PROCS[tool_id] = {"proc": proc, "started_at": time.time(), "log_lines": []}
    _capture_output(tool_id, proc)
    # Wait briefly then check if the process died immediately
    await asyncio.sleep(0.8)
    if proc.poll() is not None:
        lines = PROCS.pop(tool_id, {}).get("log_lines", [])
        error_msg = "\n".join(lines[-15:]) if lines else "Process exited immediately with no output."
        return JSONResponse({"error": f"{tool['name']} crashed on start:\n{error_msg}"}, status_code=500)
    return {"status": "started", "pid": proc.pid}


@app.post("/api/tools/{tool_id}/stop")
async def api_stop(tool_id: str):
    if tool_id not in TOOLS:
        return JSONResponse({"error": "Unknown tool"}, status_code=404)

    # Managed process: remove from registry then kill via psutil (cross-platform)
    if _is_managed_alive(tool_id):
        entry = PROCS.pop(tool_id)
        killed: list = []
        _psutil_kill(entry["proc"].pid, killed)
        await _kill_by_port(TOOLS[tool_id]["port"])
        return {"status": "stopped"}

    # Externally-started: kill only LISTENING processes (avoids killing browser clients)
    port = TOOLS[tool_id]["port"]
    killed = await _kill_by_port(port)
    if not killed:
        return {"status": "not_running"}
    return {"status": "stopped", "pids": killed}


@app.get("/api/tools/{tool_id}/logs")
async def api_logs(tool_id: str, request: Request):
    """SSE stream of log lines for a managed tool."""
    if tool_id not in TOOLS:
        return JSONResponse({"error": "Unknown tool"}, status_code=404)

    async def _gen():
        sent = 0
        while True:
            if await request.is_disconnected():
                break
            entry = PROCS.get(tool_id)
            if entry:
                lines: list = entry["log_lines"]
                while sent < len(lines):
                    yield f"data: {json.dumps(lines[sent])}\n\n"
                    sent += 1
            else:
                sent = 0
                yield f"data: {json.dumps('--- Tool not running ---')}\n\n"
                await asyncio.sleep(3)
                continue
            await asyncio.sleep(0.25)

    return StreamingResponse(
        _gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/tools/{tool_id}/stats")
async def api_stats(tool_id: str):
    if tool_id not in TOOLS:
        return JSONResponse({"error": "Unknown tool"}, status_code=404)
    managed = _is_managed_alive(tool_id)
    is_up = await _health(tool_id)
    entry = PROCS.get(tool_id)
    return {
        "id": tool_id,
        "up": is_up,
        "managed": managed,
        "uptime": int(time.time() - entry["started_at"]) if managed and entry else None,
        "pid": entry["proc"].pid if managed and entry else None,
        "log_count": len(entry["log_lines"]) if entry else 0,
    }


# ---------------------------------------------------------------------------
# Config (persists Jira credentials etc.)
# ---------------------------------------------------------------------------

@app.get("/api/config")
async def get_config():
    cfg = _load_config()
    # Never send the raw token to the browser — send a masked hint instead
    token = cfg.get("jira_api_token", "")
    return {
        "jira_base_url":   cfg.get("jira_base_url", ""),
        "jira_email":      cfg.get("jira_email", ""),
        "jira_server_type": cfg.get("jira_server_type", "cloud"),
        "jira_token_saved": bool(token),
    }

class ConfigUpdate(BaseModel):
    jira_base_url:    Optional[str] = None
    jira_email:       Optional[str] = None
    jira_api_token:   Optional[str] = None
    jira_server_type: Optional[str] = None

@app.post("/api/config")
async def save_config(body: ConfigUpdate):
    data = {k: v for k, v in body.model_dump().items() if v is not None}
    _save_config(data)
    return {"saved": True}


# ---------------------------------------------------------------------------
# Calculated Fields Extractor
# ---------------------------------------------------------------------------

def _run_extraction(twbx_path: Path, output_path: Path) -> int:
    from export_calculated_fields import export
    return export([twbx_path], str(output_path))


@app.post("/api/calc-fields/extract")
async def calc_fields_extract(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".twbx"):
        return JSONResponse({"error": "Only .twbx files are supported"}, status_code=400)

    job_id = uuid.uuid4().hex[:8]
    stem = Path(file.filename).stem

    # Save upload with original filename so the Workbook column is correct
    tmp_dir = Path(tempfile.mkdtemp())
    twbx_path = tmp_dir / file.filename
    twbx_path.write_bytes(await file.read())

    output_path = CF_OUTPUT_DIR / f"{stem}_{job_id}.xlsx"

    try:
        loop = asyncio.get_event_loop()
        count = await loop.run_in_executor(None, _run_extraction, twbx_path, output_path)
        dl_name = f"{stem} - Data Dictionary.xlsx"
        CF_JOBS[job_id] = {"status": "done", "path": str(output_path), "count": count, "filename": dl_name}
        return {"job_id": job_id, "count": count, "filename": dl_name}
    except Exception as exc:
        CF_JOBS[job_id] = {"status": "error", "error": str(exc)}
        return JSONResponse({"error": str(exc)}, status_code=500)
    finally:
        try:
            twbx_path.unlink(missing_ok=True)
            tmp_dir.rmdir()
        except Exception:
            pass


@app.get("/api/calc-fields/download/{job_id}")
async def calc_fields_download(job_id: str):
    job = CF_JOBS.get(job_id)
    if not job or job["status"] != "done":
        return JSONResponse({"error": "Job not found or not complete"}, status_code=404)
    return FileResponse(
        job["path"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=job["filename"],
    )


# ---------------------------------------------------------------------------
# Timesheet Report
# ---------------------------------------------------------------------------

class _JiraCreds(BaseModel):
    base_url: str
    email: str
    api_token: str
    server_type: str = "cloud"

class TimesheetRequest(BaseModel):
    credentials: _JiraCreds
    year: int
    month: int

class ConsolidatedRequest(BaseModel):
    credentials: _JiraCreds
    year: int


def _jira_auth(creds: _JiraCreds) -> dict:
    if creds.server_type == "server":
        auth = f"Bearer {creds.api_token}"
    else:
        auth = "Basic " + base64.b64encode(f"{creds.email}:{creds.api_token}".encode()).decode()
    return {"Authorization": auth, "Accept": "application/json", "Content-Type": "application/json"}

def _jira_api(creds: _JiraCreds) -> str:
    v = "2" if creds.server_type == "server" else "3"
    return f"{creds.base_url.rstrip('/')}/rest/api/{v}"

def _parse_jira_ts(ts: str) -> datetime:
    return datetime.strptime(ts[:19].replace("T", " "), "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)

async def _find_est_field(creds: _JiraCreds) -> Optional[str]:
    async with httpx.AsyncClient(verify=False) as client:
        r = await client.get(f"{_jira_api(creds)}/field", headers=_jira_auth(creds), timeout=15)
    if r.status_code != 200:
        return None
    for f in r.json():
        name = f["name"].lower()
        if "estimated" in name and "completion" in name:
            return f["id"]
    return None

def _pick_due(fields: dict, project_key: str, est_id: Optional[str]) -> str:
    if project_key.upper() == "BACS" and est_id:
        raw = fields.get(est_id) or ""
        return raw[:10] if raw else ""
    return fields.get("duedate") or ""

_JIRA_ACTIVITY_MAP: dict = {
    "BACS": "ONGOING_SUPPORT",
    "GISS": "SOLUTION_STNDUP",
}

# (client_name_lower, activity_code_upper) -> (project_code, activity_code)
_PROJECT_LOOKUP: dict = {}
_CLIENT_INDEX: list = []  # list of original client name strings

def _load_project_lookup() -> None:
    global _PROJECT_LOOKUP, _CLIENT_INDEX
    xlsx = BASE / "Detail.xlsx"
    if not xlsx.exists():
        return
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(str(xlsx), read_only=True, data_only=True)
        ws = wb.active
        lookup: dict = {}
        clients: set = set()
        last_client = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            client   = vals[0] if vals[0] else None
            proj_code = vals[4] if len(vals) > 4 else None
            act_code  = vals[6] if len(vals) > 6 else None
            if client:
                last_client = str(client).strip()
            if last_client and proj_code and act_code:
                key = (last_client.lower(), str(act_code).strip().upper())
                lookup[key] = (str(proj_code).strip(), str(act_code).strip())
                clients.add(last_client)
        wb.close()
        _PROJECT_LOOKUP = lookup
        _CLIENT_INDEX = list(clients)
    except Exception:
        pass

def _match_client(hint: str) -> Optional[str]:
    if not hint:
        return None
    h = hint.strip().lower()
    for name in _CLIENT_INDEX:
        if name.lower() == h:
            return name
    for name in _CLIENT_INDEX:
        if name.lower().startswith(h):
            return name
    for name in _CLIENT_INDEX:
        if h in name.lower():
            return name
    return None

def _get_project_info(summary: str, jira_project_key: str) -> tuple:
    """Returns (project_code, activity_code) looked up from Detail.xlsx."""
    client_hint = summary.split("|")[0].strip() if "|" in summary else ""
    act_code = _JIRA_ACTIVITY_MAP.get(jira_project_key.upper(), "")
    if not client_hint:
        return ("", act_code)
    matched = _match_client(client_hint)
    if not matched:
        return ("", act_code)
    info = _PROJECT_LOOKUP.get((matched.lower(), act_code.upper()))
    if info:
        return info  # (project_code, activity_code)
    return ("", act_code)

_load_project_lookup()

def _build_timesheet_sheet(ws, rows: list, month_label: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    CBRE_GREEN = "003F2D"; CBRE_MID = "005240"; EVEN = "F0F2F1"
    thin = Side(style="thin", color="C8D8D4")
    border = Border(bottom=thin, right=thin)
    HEADERS = ["Ticket #", "Category", "Summary", "Reporter", "Project Code", "Activity Code", "Due Date", "Start Date", "End Date"]
    ws.merge_cells(f"A1:{chr(64+len(HEADERS))}1")
    t = ws["A1"]; t.value = f"Timesheet — {month_label}"
    t.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    t.fill = PatternFill(start_color=CBRE_GREEN, end_color=CBRE_GREEN, fill_type="solid")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = PatternFill(start_color=CBRE_MID, end_color=CBRE_MID, fill_type="solid")
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border
    ws.row_dimensions[2].height = 22
    even_fill = PatternFill(start_color=EVEN, end_color=EVEN, fill_type="solid")
    link_font = Font(color="0563C1", underline="single", name="Calibri", size=10)
    body_font = Font(name="Calibri", size=10)
    for i, row in enumerate(rows):
        r = i + 3; fill = even_fill if i % 2 == 0 else None
        tc = ws.cell(row=r, column=1, value=row["key"]); tc.hyperlink = row["url"]
        tc.font = link_font; tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border = border
        if fill: tc.fill = fill
        for col, val, center, wrap in [
            (2, row.get("category",""), True, False),
            (3, row["summary"], False, True),
            (4, row.get("reporter",""), True, False),
            (5, row.get("project_code",""), True, False),
            (6, row.get("activity_code",""), True, False),
        ]:
            c = ws.cell(row=r, column=col, value=val); c.font = body_font
            c.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=wrap)
            c.border = border
            if fill: c.fill = fill
        for col, val in [(7, row.get("due_date","")), (8, row["start_date"]), (9, row["end_date"])]:
            dc = ws.cell(row=r, column=col, value=val if val else "—")
            if val and val != "—": dc.number_format = "MM/DD/YYYY"
            dc.alignment = Alignment(horizontal="center", vertical="center")
            dc.font = body_font; dc.border = border
            if fill: dc.fill = fill
        ws.row_dimensions[r].height = 18
    for col, w in zip("ABCDEFGHI", [15, 14, 55, 22, 18, 18, 14, 14, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    if rows: ws.auto_filter.ref = f"A2:I{len(rows)+2}"


async def _fetch_issues_and_comments(creds: _JiraCreds, jql: str, account_id: str, user_email: str, extra_field: str = ""):
    fields = f"summary,status,assignee,issuetype,reporter,duedate{','+extra_field if extra_field else ''}"
    all_issues = []; start_at = 0
    async with httpx.AsyncClient(verify=False) as client:
        while True:
            r = await client.get(f"{_jira_api(creds)}/search/jql", headers=_jira_auth(creds),
                params={"jql": jql, "fields": fields, "maxResults": 100, "startAt": start_at}, timeout=30)
            if r.status_code != 200:
                raise HTTPException(400, detail=f"Jira search error: {r.text[:300]}")
            data = r.json(); all_issues.extend(data.get("issues", []))
            if len(all_issues) >= data.get("total", 0) or not data.get("issues"): break
            start_at += 100
    sem = asyncio.Semaphore(10)
    async def _get(client, key):
        async with sem:
            r = await client.get(f"{_jira_api(creds)}/issue/{key}/comment", headers=_jira_auth(creds),
                params={"maxResults": 500, "orderBy": "created"}, timeout=30)
            return key, r
    async with httpx.AsyncClient(verify=False) as client:
        raw = await asyncio.gather(*[_get(client, i["key"]) for i in all_issues])
    return all_issues, raw


def _creds_from_request(creds: _JiraCreds) -> _JiraCreds:
    """Fill any missing credential fields from saved config."""
    cfg = _load_config()
    return _JiraCreds(
        base_url    = creds.base_url    or cfg.get("jira_base_url", ""),
        email       = creds.email       or cfg.get("jira_email", ""),
        api_token   = creds.api_token   or cfg.get("jira_api_token", ""),
        server_type = creds.server_type or cfg.get("jira_server_type", "cloud"),
    )


@app.post("/api/timesheet")
async def timesheet_monthly(req: TimesheetRequest):
    creds = _creds_from_request(req.credentials); year = req.year; month = req.month
    month_start = date(year, month, 1)
    month_end   = date(year, month, calendar.monthrange(year, month)[1])
    async with httpx.AsyncClient(verify=False) as client:
        me_r = await client.get(f"{_jira_api(creds)}/myself", headers=_jira_auth(creds), timeout=10)
    if me_r.status_code != 200:
        raise HTTPException(401, detail="Could not verify Jira credentials.")
    me = me_r.json()
    account_id = me.get("accountId") or me.get("name", "")
    user_email = me.get("emailAddress", "")
    base_url   = creds.base_url.rstrip("/")
    est_field  = await _find_est_field(creds)
    jql = (f'(assignee = currentUser() OR assignee was currentUser() OR reporter = currentUser()) '
           f'AND updated >= "{month_start}" AND updated <= "{month_end}" ORDER BY key ASC')
    all_issues, raw = await _fetch_issues_and_comments(creds, jql, account_id, user_email, est_field or "")
    if not all_issues:
        raise HTTPException(404, detail="No tickets found for this month.")
    issue_map = {i["key"]: i for i in all_issues}
    rows = []
    for key, resp in raw:
        if resp.status_code != 200: continue
        comments = resp.json().get("comments", [])
        my_dates = []
        for c in comments:
            auth = c.get("author") or {}
            if not (auth.get("accountId") == account_id or auth.get("name") == account_id or auth.get("emailAddress") == user_email):
                continue
            try:
                d = _parse_jira_ts(c["created"]).date()
                if month_start <= d <= month_end: my_dates.append(d)
            except Exception: pass
        if not my_dates: continue
        f = issue_map[key]["fields"]
        summary = f.get("summary", "")
        proj_code, act_code = _get_project_info(summary, key.split("-")[0])
        rows.append({"key": key, "summary": summary, "url": f"{base_url}/browse/{key}",
            "category": (f.get("issuetype") or {}).get("name",""),
            "reporter": (f.get("reporter") or {}).get("displayName",""),
            "project_code": proj_code, "activity_code": act_code,
            "due_date": _pick_due(f, key.split("-")[0], est_field),
            "start_date": min(my_dates), "end_date": max(my_dates)})
    if not rows:
        raise HTTPException(404, detail="No tickets found where you commented this month.")
    rows.sort(key=lambda r: r["start_date"])
    import openpyxl
    month_label = datetime(year, month, 1).strftime("%B %Y")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = month_label[:31]
    _build_timesheet_sheet(ws, rows, month_label)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"Timesheet_{datetime(year, month, 1).strftime('%B_%Y')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


@app.post("/api/timesheet/consolidated")
async def timesheet_consolidated(req: ConsolidatedRequest):
    creds = _creds_from_request(req.credentials); year = req.year
    async with httpx.AsyncClient(verify=False) as client:
        me = (await client.get(f"{_jira_api(creds)}/myself", headers=_jira_auth(creds), timeout=10)).json()
    account_id = me.get("accountId") or me.get("name", "")
    user_email = me.get("emailAddress", "")
    base_url   = creds.base_url.rstrip("/")
    est_field  = await _find_est_field(creds)
    jql = (f'(assignee = currentUser() OR assignee was currentUser() OR reporter = currentUser()) '
           f'AND updated >= "{year}-01-01" AND updated <= "{year}-12-31" ORDER BY key ASC')
    all_issues, raw = await _fetch_issues_and_comments(creds, jql, account_id, user_email, est_field or "")
    if not all_issues:
        raise HTTPException(404, detail=f"No tickets found for {year}.")
    issue_map = {i["key"]: i for i in all_issues}
    month_data: dict = {}
    for key, resp in raw:
        if resp.status_code != 200: continue
        month_dates: dict = {}
        for c in resp.json().get("comments", []):
            auth = c.get("author") or {}
            if not (auth.get("accountId") == account_id or auth.get("name") == account_id or auth.get("emailAddress") == user_email):
                continue
            try:
                d = _parse_jira_ts(c["created"]).date()
                if d.year == year: month_dates.setdefault(d.month, []).append(d)
            except Exception: pass
        if not month_dates: continue
        f = issue_map[key]["fields"]
        for m, dates in month_dates.items():
            summary = f.get("summary", "")
            proj_code, act_code = _get_project_info(summary, key.split("-")[0])
            month_data.setdefault(m, []).append({"key": key, "summary": summary,
                "url": f"{base_url}/browse/{key}", "category": (f.get("issuetype") or {}).get("name",""),
                "reporter": (f.get("reporter") or {}).get("displayName",""),
                "project_code": proj_code, "activity_code": act_code,
                "due_date": _pick_due(f, key.split("-")[0], est_field),
                "start_date": min(dates), "end_date": max(dates)})
    if not month_data:
        raise HTTPException(404, detail=f"No comment activity found for {year}.")
    for m in month_data: month_data[m].sort(key=lambda r: r["start_date"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    CBRE_GREEN = "003F2D"; CBRE_MID = "005240"
    thin = Side(style="thin", color="C8D8D4"); border = Border(bottom=thin, right=thin)
    wb = openpyxl.Workbook(); ws_all = wb.active; ws_all.title = "All Months"
    ALL_HDRS = ["Month","Ticket #","Category","Summary","Reporter","Project Code","Activity Code","Due Date","Start Date","End Date"]
    ws_all.merge_cells(f"A1:{chr(64+len(ALL_HDRS))}1")
    t = ws_all["A1"]; t.value = f"Consolidated Timesheet — {year}"
    t.font = Font(bold=True, size=15, color="FFFFFF", name="Calibri")
    t.fill = PatternFill(start_color=CBRE_GREEN, end_color=CBRE_GREEN, fill_type="solid")
    t.alignment = Alignment(horizontal="center", vertical="center"); ws_all.row_dimensions[1].height = 40
    for col, h in enumerate(ALL_HDRS, 1):
        c = ws_all.cell(row=2, column=col, value=h)
        c.fill = PatternFill(start_color=CBRE_MID, end_color=CBRE_MID, fill_type="solid")
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border
    ws_all.row_dimensions[2].height = 22
    MONTH_COLORS = ["E8F5E9","FFF8E1","E3F2FD","FCE4EC","F3E5F5","E0F7FA","FFF3E0","E8EAF6","F1F8E9","FBE9E7","E0F2F1","F9FBE7"]
    link_font = Font(color="0563C1", underline="single", name="Calibri", size=10)
    body_font = Font(name="Calibri", size=10); bold_font = Font(bold=True, name="Calibri", size=10)
    sr = 3
    for m in sorted(month_data.keys()):
        month_label = datetime(year, m, 1).strftime("%B %Y")
        color = MONTH_COLORS[(m-1)%12]
        mfill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for row in month_data[m]:
            def _c(col, val, *, link=False, center=True, wrap=False, date_fmt=False):
                cell = ws_all.cell(row=sr, column=col, value=val if val else "—")
                cell.font = link_font if link else body_font; cell.fill = mfill; cell.border = border
                cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=wrap)
                if date_fmt and val and val != "—": cell.number_format = "MM/DD/YYYY"
                if link: cell.hyperlink = row["url"]
            ws_all.cell(row=sr, column=1, value=month_label).font = bold_font
            ws_all.cell(row=sr, column=1).fill = mfill; ws_all.cell(row=sr, column=1).border = border
            ws_all.cell(row=sr, column=1).alignment = Alignment(horizontal="center", vertical="center")
            _c(2, row["key"], link=True); _c(3, row.get("category","")); _c(4, row["summary"], center=False, wrap=True)
            _c(5, row.get("reporter","")); _c(6, row.get("project_code","")); _c(7, row.get("activity_code",""))
            _c(8, row.get("due_date",""), date_fmt=True)
            _c(9, row["start_date"], date_fmt=True); _c(10, row["end_date"], date_fmt=True)
            ws_all.row_dimensions[sr].height = 18; sr += 1
    for col, w in zip("ABCDEFGHIJ", [16,15,14,55,22,18,18,14,14,14]): ws_all.column_dimensions[col].width = w
    ws_all.freeze_panes = "A3"; ws_all.auto_filter.ref = f"A2:J{sr-1}"
    for m in sorted(month_data.keys()):
        ws_m = wb.create_sheet(title=datetime(year, m, 1).strftime("%B %Y")[:31])
        _build_timesheet_sheet(ws_m, month_data[m], datetime(year, m, 1).strftime("%B %Y"))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"Timesheet_Consolidated_{year}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("monitor:app", host="0.0.0.0", port=9000, reload=False)
