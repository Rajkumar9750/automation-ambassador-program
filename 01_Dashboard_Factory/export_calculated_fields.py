"""
Export calculated fields from one or more Tableau .twbx workbooks to Excel.

Usage:
    python3 export_calculated_fields.py workbook.twbx [workbook2.twbx ...] [-o output.xlsx]
    python3 export_calculated_fields.py uploads/                           [-o output.xlsx]
"""
import os
import re
import sys
import zipfile
import tempfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import List, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import FilterColumn, Filters, CustomFilter, CustomFilters
from datetime import datetime


# ── CBRE Brand Palette ───────────────────────────────────────────────────────

C_GREEN_DARK  = "003F2D"   # CBRE primary dark green  — title / header bg
C_GREEN_MID   = "005240"   # CBRE mid green           — sub-header accent
C_GREEN_LIGHT = "EAF2EF"   # very light green         — alt row tint
C_BORDER      = "B2CCCA"   # muted green-grey         — cell borders
C_WHITE       = "FFFFFF"
C_TEXT_DARK   = "1A2E25"   # near-black green tint    — body text
C_MUTED       = "6B8F86"   # muted green              — secondary text
C_HIDDEN_YES  = "FFF0F0"   # soft red tint            — hidden = Yes rows
C_HIDDEN_NO   = "F0FFF4"   # soft green tint          — hidden = No rows
C_BADGE_CALC  = "E8F5E9"   # field type badge bg
C_BADGE_TEXT  = "2E7D32"   # field type badge text

# Fills
TITLE_FILL    = PatternFill("solid", fgColor=C_GREEN_DARK)
SUBHDR_FILL   = PatternFill("solid", fgColor=C_GREEN_MID)
HEADER_FILL   = PatternFill("solid", fgColor=C_GREEN_DARK)
ALT_FILL      = PatternFill("solid", fgColor=C_GREEN_LIGHT)
WHITE_FILL    = PatternFill("solid", fgColor=C_WHITE)
HIDDEN_FILL   = PatternFill("solid", fgColor=C_HIDDEN_YES)

# Borders
_THIN  = Side(style="thin",   color=C_BORDER)
_THICK = Side(style="medium", color=C_GREEN_MID)
HDR_BORDER  = Border(left=_THIN, right=_THIN, top=_THICK, bottom=_THICK)
CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN,  bottom=_THIN)

# Fonts
TITLE_FONT   = Font(bold=True,  color=C_WHITE,     size=15, name="Calibri")
SUBHDR_FONT  = Font(italic=True,color="D4EBE4",    size=10, name="Calibri")
HEADER_FONT  = Font(bold=True,  color=C_WHITE,     size=10, name="Calibri")
BODY_FONT    = Font(           color=C_TEXT_DARK,  size=10, name="Calibri")
FORMULA_FONT = Font(           color="1B4332",     size=9,  name="Consolas")
MUTED_FONT   = Font(italic=True,color=C_MUTED,    size=9,  name="Calibri")
BADGE_FONT   = Font(bold=True,  color=C_BADGE_TEXT,size=9,  name="Calibri")
LINK_FONT    = Font(           color="0563C1",     size=10, name="Calibri",
                    underline="single")

# Alignments
CENTER  = Alignment(horizontal="center", vertical="center")
LEFT    = Alignment(horizontal="left",   vertical="center")
WRAP    = Alignment(horizontal="left",   vertical="top", wrap_text=True)
TOP_CTR = Alignment(horizontal="center", vertical="top")

COLUMNS = [
    ("Workbook",           24),
    ("Datasource",         24),
    ("Field Name",         28),
    ("Original Name",      32),
    ("Formula",            58),
    ("Field Type",         15),
    ("Data Type",          12),
    ("Hidden",             10),
    ("# of Times Used",    14),
    ("Worksheets Used On", 48),
]

_N_COLS = len(COLUMNS)
_LAST_COL = get_column_letter(_N_COLS)


# ── Calculated-field parsing ─────────────────────────────────────────────────

def _resolve_formula(formula: str, name_map: Dict[str, str]) -> str:
    """Replace internal [Calculation_XXX] IDs in formula with human-readable captions."""
    def replacer(m):
        token = m.group(0)
        return f"[{name_map[token]}]" if token in name_map else token
    return re.sub(r'\[[^\]]+\]', replacer, formula)


def _extract_calc_fields(twbx_path: str) -> List[Dict]:
    """
    Parse a .twbx and return all calculated fields across all datasources,
    including original name, worksheet usage count, and worksheet list.
    """
    fields = []
    with tempfile.TemporaryDirectory() as tmp:
        with zipfile.ZipFile(twbx_path, "r") as z:
            twb_files = [f for f in z.namelist() if f.endswith(".twb")]
            if not twb_files:
                return fields
            z.extract(twb_files[0], tmp)
            twb_path = os.path.join(tmp, twb_files[0])

        with open(twb_path, encoding="utf-8", errors="replace") as f:
            content = f.read()

    try:
        root = ET.fromstring(content)
    except ET.ParseError:
        return fields

    ds_container = root.find("datasources")
    if ds_container is None:
        return fields

    # ── Build worksheet usage map: internal_col_name → [worksheet names] ──
    ws_usage: Dict[str, List[str]] = {}
    worksheets_el = root.find("worksheets")
    if worksheets_el is not None:
        for worksheet in worksheets_el.findall("worksheet"):
            ws_name = worksheet.get("name", "")
            for col in worksheet.iter("column"):
                col_name = col.get("name", "")
                if col_name:
                    ws_usage.setdefault(col_name, [])
                    if ws_name not in ws_usage[col_name]:
                        ws_usage[col_name].append(ws_name)

    # ── Extract ALL fields from each datasource ───────────────────────────
    for ds in ds_container.findall("datasource"):
        ds_name    = ds.get("name", "")
        ds_caption = ds.get("caption", ds_name)

        if ds_name == "Parameters" or not ds_name:
            continue

        name_map: Dict[str, str] = {}
        for col in ds.findall("column"):
            col_name = col.get("name", "")
            caption  = col.get("caption", "")
            if col_name and caption:
                name_map[col_name] = caption

        for col in ds.findall("column"):
            internal_name = col.get("name", "")
            caption       = col.get("caption", internal_name.strip("[]"))
            calc          = col.find("calculation")
            used_on       = ws_usage.get(internal_name, [])

            if calc is not None and calc.get("class") == "tableau":
                field_type = "Calculated Field"
                formula    = _resolve_formula(calc.get("formula", ""), name_map)
            elif calc is not None:
                field_type = "Group / Bin"
                formula    = calc.get("formula", "")
            else:
                field_type = "Database Field"
                formula    = ""

            fields.append({
                "datasource":    ds_caption or ds_name,
                "caption":       caption,
                "original_name": internal_name,
                "formula":       formula,
                "field_type":    field_type,
                "datatype":      col.get("datatype", ""),
                "hidden":        col.get("hidden", "false").lower() == "true",
                "times_used":    len(used_on),
                "worksheets":    str(used_on) if used_on else "",
            })

    return fields


# ── Excel helpers ────────────────────────────────────────────────────────────

def _title_block(ws, workbook_name: str, total: int, generated: str) -> None:
    """Rows 1-2: metadata strip + column headers."""
    # Row 1 — metadata strip
    ws.merge_cells(f"A1:{_LAST_COL}1")
    m = ws["A1"]
    m.value     = f"  {workbook_name}   ·   Generated {generated}   ·   {total} field(s) extracted"
    m.font      = SUBHDR_FONT
    m.fill      = PatternFill("solid", fgColor=C_GREEN_DARK)
    m.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Row 2 — column headers
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=2, column=col_idx, value=header)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER
        c.border    = HDR_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 24

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{_LAST_COL}2"

    # Default filter: Hidden = "No" (col index 7, 0-based)
    hidden_col_idx = next(i for i, (h, _) in enumerate(COLUMNS) if h == "Hidden")
    fc_hidden = FilterColumn(colId=hidden_col_idx)
    fc_hidden.filters = Filters(filter=["No"])
    ws.auto_filter.filterColumn.append(fc_hidden)

    # Default filter: # of Times Used ≠ 0 (deselect 0)
    used_col_idx = next(i for i, (h, _) in enumerate(COLUMNS) if h == "# of Times Used")
    fc_used = FilterColumn(colId=used_col_idx)
    fc_used.customFilters = CustomFilters(
        customFilter=[CustomFilter(operator="notEqual", val="0")]
    )
    ws.auto_filter.filterColumn.append(fc_used)


def _write_data_row(ws, row_num: int, field: dict, workbook_name: str, alt: bool) -> None:
    is_hidden = field["hidden"]
    base_fill = HIDDEN_FILL if is_hidden else (ALT_FILL if alt else WHITE_FILL)

    values = [
        workbook_name,
        field["datasource"],
        field["caption"],
        field["original_name"],
        field["formula"],
        field["field_type"],
        field["datatype"],
        "Yes" if is_hidden else "No",
        field["times_used"],
        field["worksheets"],
    ]

    for col_idx, val in enumerate(values, start=1):
        c = ws.cell(row=row_num, column=col_idx, value=val)
        c.border = CELL_BORDER
        c.fill   = base_fill

        if col_idx == 5:          # Formula — monospace, wrapped
            c.font      = FORMULA_FONT
            c.alignment = WRAP
        elif col_idx == 8:        # Hidden — color-coded badge style
            c.font      = Font(bold=True, color="C62828" if is_hidden else "2E7D32",
                               size=10, name="Calibri")
            c.alignment = CENTER
        elif col_idx == 9:        # # of Times Used — centered number
            c.font      = Font(bold=True, color=C_TEXT_DARK, size=10, name="Calibri")
            c.alignment = CENTER
        elif col_idx in (1, 2, 3, 6, 7):  # Short text columns — centered
            c.font      = BODY_FONT
            c.alignment = CENTER
        elif col_idx == 4:        # Original Name — muted, wrapped
            c.font      = MUTED_FONT
            c.alignment = WRAP
        else:                     # Worksheets Used On — wrapped left
            c.font      = BODY_FONT
            c.alignment = WRAP

    ws.row_dimensions[row_num].height = max(
        18, min(80, 18 + field["formula"].count("\n") * 14)
    )


# ── Main export ──────────────────────────────────────────────────────────────

def export(twbx_paths: List[Path], output_path: str) -> int:
    all_fields = []
    workbook_map = []

    for twbx_path in twbx_paths:
        try:
            fields = _extract_calc_fields(str(twbx_path))
        except Exception as exc:
            print(f"[ERROR] {twbx_path.name}: {exc}", file=sys.stderr)
            continue
        for f in fields:
            all_fields.append((twbx_path.stem, f))

    total     = len(all_fields)
    generated = datetime.now().strftime("%B %d, %Y  %I:%M %p")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Dictionary"
    ws.sheet_view.showGridLines = False

    # Title uses first workbook name (or generic)
    wb_label = twbx_paths[0].stem if twbx_paths else "Workbook"
    _title_block(ws, wb_label, total, generated)

    for i, (wb_name, field) in enumerate(all_fields):
        row_num = i + 3
        _write_data_row(ws, row_num, field, wb_name, alt=(i % 2 == 0))
        # Hide rows that don't match default filters: Hidden=No and Times Used > 0
        if field["hidden"] or field["times_used"] == 0:
            ws.row_dimensions[row_num].hidden = True

    # Summary row at bottom
    last_row = total + 3
    ws.merge_cells(f"A{last_row}:{_LAST_COL}{last_row}")
    sr = ws[f"A{last_row}"]
    hidden_count = sum(1 for _, f in all_fields if f["hidden"])
    calc_count = sum(1 for _, f in all_fields if f["field_type"] == "Calculated Field")
    db_count   = sum(1 for _, f in all_fields if f["field_type"] == "Database Field")
    sr.value   = f"  Total: {total} fields   ·   {calc_count} calculated   ·   {db_count} database   ·   {hidden_count} hidden"
    sr.font      = Font(italic=True, color=C_MUTED, size=9, name="Calibri")
    sr.fill      = PatternFill("solid", fgColor="F4F8F7")
    sr.alignment = Alignment(horizontal="left", vertical="center")
    sr.border    = Border(top=Side(style="medium", color=C_GREEN_MID))
    ws.row_dimensions[last_row].height = 20

    wb.save(output_path)
    return total


def _resolve_paths(inputs: list) -> List[Path]:
    paths = []
    for inp in inputs:
        p = Path(inp)
        if p.is_dir():
            paths.extend(sorted(p.glob("**/*.twbx")))
        elif p.suffix.lower() == ".twbx" and p.exists():
            paths.append(p)
        else:
            print(f"[WARN] Skipping '{inp}' — not a .twbx file or directory", file=sys.stderr)
    return paths


DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")


def main():
    file_path = input("Enter the path to the .twbx file (or folder): ").strip().strip("'\"")

    paths = _resolve_paths([file_path])
    if not paths:
        print("No .twbx files found at the given path.", file=sys.stderr)
        sys.exit(1)

    print(f"Processing {len(paths)} workbook(s)...")
    grand_total = 0
    for p in paths:
        out = os.path.join(DESKTOP, f"{p.stem}.xlsx")
        total = export([p], out)
        grand_total += total
        print(f"  {p.name} → {out} ({total} field(s))")
    print(f"Done — {grand_total} calculated field(s) exported.")


if __name__ == "__main__":
    main()
