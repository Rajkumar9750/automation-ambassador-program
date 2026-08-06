import os
import re
import sys
import requests
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

JIRA_BASE_URL = os.getenv("JIRA_BASE_URL", "").rstrip("/")

STATUS_MAP = {
    "GISS": {"start": "Implementation", "end": "Internal Review"},
    "BACS": {"start": "CS Implementation", "end": "Customer Review"},
}


def parse_dt(s):
    # Normalize Z and ±HHMM (no colon) to ±HH:MM for Python 3.9 compatibility
    s = s.replace("Z", "+00:00")
    s = re.sub(r"([+-])(\d{2})(\d{2})$", r"\1\2:\3", s)
    return datetime.fromisoformat(s)


def check_env():
    if not os.getenv("JIRA_BASE_URL"):
        raise EnvironmentError("Missing JIRA_BASE_URL in .env")


def get_current_user(base_url, auth):
    response = requests.get(
        f"{base_url}/rest/api/3/myself",
        auth=auth,
        headers={"Accept": "application/json"},
    )
    response.raise_for_status()
    return response.json()


def get_issues(jql, base_url, auth):
    issues = []
    next_page_token = None
    max_results = 100

    while True:
        params = {
            "jql": jql,
            "maxResults": max_results,
            "fields": "summary,status,created,reporter,customfield_12536,customfield_12583,customfield_12531,customfield_12532",
        }
        if next_page_token:
            params["nextPageToken"] = next_page_token

        response = requests.get(
            f"{base_url}/rest/api/3/search/jql",
            auth=auth,
            headers={"Accept": "application/json"},
            params=params,
        )
        response.raise_for_status()
        data = response.json()

        batch = data.get("issues", [])
        issues.extend(batch)

        next_page_token = data.get("nextPageToken")
        if not next_page_token or len(batch) < max_results:
            break

    return issues


def get_changelog(issue_key, base_url, auth):
    histories = []
    start_at = 0
    max_results = 100

    while True:
        response = requests.get(
            f"{base_url}/rest/api/3/issue/{issue_key}/changelog",
            auth=auth,
            headers={"Accept": "application/json"},
            params={"startAt": start_at, "maxResults": max_results},
        )
        response.raise_for_status()
        data = response.json()

        histories.extend(data.get("values", []))
        start_at += max_results

        if start_at >= data.get("total", 0):
            break

    return histories


def get_work_periods(issue_key, category, window_start, window_end, user_account_id, base_url, auth):
    """
    Returns a list of (start_date, end_date) tuples — one per work cycle.
    A ticket bounced back in the same month produces multiple entries.
    """
    start_status = STATUS_MAP[category]["start"].lower()

    status_changes   = []
    assignee_changes = []

    for history in get_changelog(issue_key, base_url, auth):
        created   = history.get("created", "")
        author_id = (history.get("author") or {}).get("accountId", "")
        dt        = parse_dt(created)
        for item in history.get("items", []):
            field = item.get("field")
            if field == "status":
                status_changes.append({
                    "dt":        dt,
                    "frm":       item.get("fromString", "").lower(),
                    "to":        item.get("toString", "").lower(),
                    "author_id": author_id,
                })
            elif field == "assignee":
                assignee_changes.append({
                    "dt":      dt,
                    "to_id":   item.get("to", ""),
                    "from_id": item.get("from", ""),
                })

    status_changes.sort(key=lambda x: x["dt"])
    assignee_changes.sort(key=lambda x: x["dt"])

    initial_assignee = assignee_changes[0]["from_id"] if assignee_changes else None

    def assignee_at(target_dt):
        current = initial_assignee
        for ac in assignee_changes:
            if ac["dt"] <= target_dt:
                current = ac["to_id"]
            else:
                break
        return current

    def find_exit(after_dt):
        for t in status_changes:
            if t["frm"] == start_status and t["dt"] > after_dt:
                return t["dt"], t["author_id"]
        return None, None

    def find_reassign_away(after_dt):
        for ac in assignee_changes:
            if ac["from_id"] == user_account_id and ac["dt"] > after_dt:
                return ac["dt"]
        return None

    def resolve_end(entry_dt, exit_dt, exit_author, reassign_dt):
        # End date = whichever comes first: ticket leaving the start status,
        # or the user being reassigned away. Reassignment always means work ended.
        if exit_dt and reassign_dt:
            return min(exit_dt, reassign_dt)
        return exit_dt or reassign_dt

    periods = []
    seen = set()

    # Primary: all start-status entries in window where user was involved
    for t in status_changes:
        if t["to"] != start_status:
            continue
        if not (window_start <= t["dt"].date() <= window_end):
            continue
        if t["dt"] in seen:
            continue

        is_author   = t["author_id"] == user_account_id
        is_assignee = assignee_at(t["dt"]) == user_account_id

        exit_dt, exit_author = find_exit(t["dt"])

        # Include if user authored the entry, was the assignee, or authored the exit
        if not (is_author or is_assignee or exit_author == user_account_id):
            continue

        seen.add(t["dt"])
        reassign_dt = find_reassign_away(t["dt"])
        end_dt = resolve_end(t["dt"], exit_dt, exit_author, reassign_dt)
        periods.append((t["dt"], end_dt))

    # Fallback: user assigned to ticket already in start status
    if not periods:
        for ac in assignee_changes:
            if ac["to_id"] == user_account_id and window_start <= ac["dt"].date() <= window_end:
                status_then = None
                for sc in status_changes:
                    if sc["dt"] <= ac["dt"]:
                        status_then = sc["to"]
                if status_then == start_status:
                    exit_dt, exit_author = find_exit(ac["dt"])
                    reassign_dt = find_reassign_away(ac["dt"])
                    end_dt = resolve_end(ac["dt"], exit_dt, exit_author, reassign_dt)
                    periods.append((ac["dt"], end_dt))

    return periods


def fetch_tickets(month, year, base_url, auth):
    from calendar import monthrange
    import datetime as dt

    rows = []
    errors = []

    last_day = monthrange(year, month)[1]

    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1

    prev_last_day   = monthrange(prev_year, prev_month)[1]
    window_start_dy = max(prev_last_day - 2, 1)

    date_from      = f"{prev_year}-{prev_month:02d}-{window_start_dy:02d}"
    date_to        = f"{year}-{month:02d}-{last_day:02d} 23:59"
    window_start   = dt.date(prev_year, prev_month, window_start_dy)
    window_end     = dt.date(year, month, last_day)
    first_of_month = dt.date(year, month, 1)

    user_account_id = get_current_user(base_url, auth)["accountId"]

    for category in ["GISS", "BACS"]:
        start_status = STATUS_MAP[category]["start"]
        try:
            jql = (
                f'project = {category} AND '
                f'assignee was currentUser() AND '
                f'status changed to "{start_status}" '
                f'during ("{date_from}", "{date_to}")'
            )
            issues = get_issues(jql, base_url, auth)

            for issue in issues:
                periods = get_work_periods(
                    issue["key"], category, window_start, window_end,
                    user_account_id, base_url, auth
                )

                f             = issue["fields"]
                reporter      = (f.get("reporter") or {}).get("displayName", "—")
                project_code  = f.get("customfield_12531") or "—"
                activity_code = "SOLUTION_STANDUP" if category == "GISS" else (f.get("customfield_12532") or "—")
                due_date      = f.get("customfield_12536") or "—" if category == "GISS" else f.get("customfield_12583") or "—"

                for start_date, end_date in periods:
                    if not start_date:
                        continue
                    if end_date and end_date.date() < first_of_month:
                        continue
                    # Cap start/end dates to the selected month boundaries
                    last_of_month  = dt.date(year, month, last_day)
                    display_start  = max(start_date.date(), first_of_month)
                    display_end    = min(end_date.date(), last_of_month) if end_date else None
                    duration = (
                        (display_end - display_start).days + 1 if display_end else None
                    )
                    end_str = display_end.strftime("%Y-%m-%d") if display_end else "In Progress"

                    rows.append({
                        "ticketId":    issue["key"],
                        "summary":     f.get("summary", ""),
                        "category":    category,
                        "reporter":    reporter,
                        "projectCode": project_code,
                        "activityCode": activity_code,
                        "dueDate":     due_date,
                        "startDate":   display_start.strftime("%Y-%m-%d"),
                        "endDate":     end_str,
                        "duration":    duration if duration is not None else "N/A",
                    })
        except requests.HTTPError as e:
            errors.append(f"{category}: {e}")

    rows.sort(key=lambda r: r["startDate"])
    return rows, errors


TICKET_PALETTE = [
    "DBEAFE", "DCFCE7", "F3E8FF", "FFEDD5", "FCE7F3",
    "CCFBF1", "FEF9C3", "E0E7FF", "FFE4E6", "E0F2FE",
    "ECFCCB", "FEF3C7", "EDE9FE", "CFFAFE", "D1FAE5",
    "FEE2E2", "FDF2F8", "F0FDF4", "EFF6FF", "FFFBEB",
]


def export_to_excel(rows, month, year):
    import io as _io
    month_str = datetime(year, month, 1).strftime("%B_%Y")
    filename  = f"Timesheet_{month_str}.xlsx"

    # Assign a unique colour to each ticket ID
    colour_map = {}
    for row in rows:
        tid = row["ticketId"]
        if tid not in colour_map:
            colour_map[tid] = TICKET_PALETTE[len(colour_map) % len(TICKET_PALETTE)]

    wb = Workbook()
    ws = wb.active
    ws.title = datetime(year, month, 1).strftime("%b %Y")

    # ── Styles ────────────────────────────────────────────────
    thin        = Side(style="thin",   color="D1D5DB")
    thick_left  = Side(style="medium", color="9CA3AF")
    no_side     = Side(style=None)

    def row_border(left_color):
        left = Side(style="medium", color=left_color)
        return Border(left=left, right=no_side, top=no_side, bottom=Side(style="thin", color="E5E7EB"))

    header_font  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    header_fill  = PatternFill(start_color="002147", end_color="002147", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    header_border= Border(bottom=Side(style="medium", color="FFFFFF"))

    data_font    = Font(size=10, name="Calibri", color="1C2536")
    link_font    = Font(size=10, name="Calibri", color="0047AB", underline="single", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    columns  = ["Ticket ID", "Summary", "Category", "Reporter",
                "Project Code", "Activity Code", "Due Date",
                "Start Date", "End Date", "Days"]
    col_keys = ["ticketId", "summary", "category", "reporter",
                "projectCode", "activityCode", "dueDate",
                "startDate", "endDate", "duration"]

    # ── Column widths ─────────────────────────────────────────
    col_widths = [14, 70, 10, 22, 16, 18, 12, 12, 12, 6]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Header row ────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = header_border

    # ── Data rows ─────────────────────────────────────────────
    for row_idx, row in enumerate(rows, 2):
        hex_colour = colour_map[row["ticketId"]]
        bg_fill    = PatternFill(start_color=hex_colour, end_color=hex_colour, fill_type="solid")
        b          = row_border(hex_colour)

        for col_idx, key in enumerate(col_keys, 1):
            val   = row[key]
            cell  = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = bg_fill
            cell.border    = b
            cell.alignment = center_align if key in ("category", "duration", "startDate", "endDate", "dueDate") else left_align

            if key == "ticketId":
                cell.hyperlink = f"{JIRA_BASE_URL}/browse/{row[key]}"
                cell.font = link_font
            else:
                cell.font = data_font

    # ── Auto-filter & freeze ──────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    ws.freeze_panes   = "A2"

    # ── Daily View tab ────────────────────────────────────────
    _build_daily_sheet(wb, rows, month, year, colour_map, JIRA_BASE_URL)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Also save to disk for CLI usage
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    with open(output_path, "wb") as f:
        f.write(buf.getvalue())
    buf.seek(0)
    return buf, filename


def _build_daily_sheet(wb, rows, month, year, colour_map, jira_base_url):
    import datetime as dt
    from calendar import monthrange

    last_day = monthrange(year, month)[1]
    first    = dt.date(year, month, 1)
    last     = dt.date(year, month, last_day)

    # Expand each ticket's date range into individual days within the month
    day_map: dict = {}
    for row in rows:
        try:
            start = dt.date.fromisoformat(row["startDate"])
        except Exception:
            continue
        try:
            end = dt.date.fromisoformat(row["endDate"])
        except Exception:
            end = last
        start = max(start, first)
        end   = min(end,   last)
        d = start
        while d <= end:
            if d.weekday() < 5:  # 0=Mon … 4=Fri, skip Sat/Sun
                day_map.setdefault(d, []).append(row)
            d += dt.timedelta(days=1)

    ws = wb.create_sheet(title="Daily View")

    # ── Styles ───────────────────────────────────────────────
    thin   = Side(style="thin",   color="D1D5DB")
    medium = Side(style="medium", color="9CA3AF")
    full_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font   = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    header_fill   = PatternFill(start_color="002147", end_color="002147", fill_type="solid")
    header_align  = Alignment(horizontal="center", vertical="center")
    header_border = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="FFFFFF"))
    data_font     = Font(size=10, name="Calibri", color="1C2536")
    bold_font     = Font(size=10, name="Calibri", color="1C2536", bold=True)
    link_font     = Font(size=10, name="Calibri", color="0047AB", underline="single", bold=True)
    center_align  = Alignment(horizontal="center", vertical="center")
    left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    date_align    = Alignment(horizontal="center", vertical="center", wrap_text=False)

    COLUMNS = ["Date", "Ticket ID", "Project Code", "Activity Code", "Summary"]
    WIDTHS  = [14,      14,          16,              18,              70]

    ws.row_dimensions[1].height = 28
    for ci, (name, w) in enumerate(zip(COLUMNS, WIDTHS), 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = header_font; c.fill = header_fill
        c.alignment = header_align; c.border = header_border
        ws.column_dimensions[get_column_letter(ci)].width = w

    DAY_BANDS = ["EFF6FF", "F0FDF4"]  # alternating blue/green tint per day
    ri = 2

    for day in sorted(day_map.keys()):
        tickets   = day_map[day]
        n         = len(tickets)
        band_hex  = DAY_BANDS[day.day % 2]
        band_fill = PatternFill(start_color=band_hex, end_color=band_hex, fill_type="solid")
        date_str  = day.strftime("%a, %d %b")
        row_start = ri

        for i, row in enumerate(tickets):
            tk_hex  = colour_map[row["ticketId"]]
            tk_fill = PatternFill(start_color=tk_hex, end_color=tk_hex, fill_type="solid")

            # Ticket ID
            tc = ws.cell(row=ri, column=2, value=row["ticketId"])
            tc.hyperlink = f"{jira_base_url}/browse/{row['ticketId']}"
            tc.font = link_font; tc.fill = tk_fill
            tc.alignment = center_align; tc.border = full_border

            # Project Code
            pc = ws.cell(row=ri, column=3, value=row.get("projectCode", ""))
            pc.font = data_font; pc.fill = tk_fill
            pc.alignment = center_align; pc.border = full_border

            # Activity Code
            ac = ws.cell(row=ri, column=4, value=row.get("activityCode", ""))
            ac.font = data_font; ac.fill = tk_fill
            ac.alignment = center_align; ac.border = full_border

            # Summary
            sc = ws.cell(row=ri, column=5, value=row.get("summary", ""))
            sc.font = data_font; sc.fill = tk_fill
            sc.alignment = left_align; sc.border = full_border

            ws.row_dimensions[ri].height = 18
            ri += 1

        # Merge the date cell across all ticket rows for this day and set
        # borders on every cell in the range (openpyxl requirement for merged cells)
        row_end = ri - 1
        if n > 1:
            ws.merge_cells(start_row=row_start, start_column=1,
                           end_row=row_end,     end_column=1)
        for r in range(row_start, row_end + 1):
            top_side    = thin if r == row_start else Side(style=None)
            bottom_side = thin if r == row_end   else Side(style=None)
            ws.cell(row=r, column=1).border = Border(
                left=thin, right=thin, top=top_side, bottom=bottom_side
            )
            ws.cell(row=r, column=1).fill = band_fill
        date_cell = ws.cell(row=row_start, column=1, value=date_str)
        date_cell.font      = bold_font
        date_cell.alignment = date_align

    ws.freeze_panes = "A2"
    if ri > 2:
        ws.auto_filter.ref = f"A1:E{ri - 1}"


MONTH_BAND_COLORS = [
    "E8F5E9", "FFF8E1", "E3F2FD", "FCE4EC", "F3E5F5", "E0F7FA",
    "FFF3E0", "E8EAF6", "F1F8E9", "FBE9E7", "E0F2F1", "F9FBE7",
]


def export_year_to_excel(month_data: dict, year: int):
    """
    month_data: {month_int: [row_dicts]}
    Produces one workbook with:
      - "All Months" consolidated sheet (month band + ticket colour)
      - One sheet per month that has data
    """
    filename    = f"Timesheet_{year}.xlsx"
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

    # Build a global colour map across every row
    colour_map: dict = {}
    for rows in month_data.values():
        for row in rows:
            tid = row["ticketId"]
            if tid not in colour_map:
                colour_map[tid] = TICKET_PALETTE[len(colour_map) % len(TICKET_PALETTE)]

    no_side = Side(style=None)

    def row_border(left_color: str):
        return Border(
            left=Side(style="medium", color=left_color),
            right=no_side, top=no_side,
            bottom=Side(style="thin", color="E5E7EB"),
        )

    header_font   = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    header_fill   = PatternFill(start_color="003F2D", end_color="003F2D", fill_type="solid")
    header_align  = Alignment(horizontal="center", vertical="center")
    header_border = Border(bottom=Side(style="medium", color="FFFFFF"))
    data_font     = Font(size=10, name="Calibri", color="1C2536")
    bold_font     = Font(size=10, name="Calibri", color="1C2536", bold=True)
    link_font     = Font(size=10, name="Calibri", color="003F2D", underline="single", bold=True)
    center_align  = Alignment(horizontal="center", vertical="center")
    left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    COLS      = ["Ticket ID", "Summary", "Category", "Reporter",
                 "Project Code", "Activity Code", "Due Date", "Start Date", "End Date", "Days"]
    COL_KEYS  = ["ticketId", "summary", "category", "reporter",
                 "projectCode", "activityCode", "dueDate", "startDate", "endDate", "duration"]
    COL_W     = [14, 70, 10, 22, 16, 18, 12, 12, 12, 6]

    ALL_COLS  = ["Month"] + COLS
    ALL_KEYS  = ["_month_label"] + COL_KEYS
    ALL_W     = [14] + COL_W

    def _write_header(ws, columns, widths):
        ws.row_dimensions[1].height = 30
        for ci, name in enumerate(columns, 1):
            c = ws.cell(row=1, column=ci, value=name)
            c.font = header_font; c.fill = header_fill
            c.alignment = header_align; c.border = header_border
            ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]

    def _write_row(ws, ri, row, keys, bg_hex, border, month_band_fill=None):
        bg_fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
        for ci, key in enumerate(keys, 1):
            val  = row.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            # First column of All-Months sheet uses the month band colour
            cell.fill = (month_band_fill if (month_band_fill and ci == 1) else bg_fill)
            cell.border = border
            cell.alignment = (
                center_align if key in ("category", "duration", "startDate",
                                        "endDate", "dueDate", "_month_label")
                else left_align
            )
            if key == "ticketId":
                cell.hyperlink = f"{JIRA_BASE_URL}/browse/{val}"
                cell.font = link_font
            elif key == "_month_label":
                cell.font = bold_font
            else:
                cell.font = data_font

    wb = Workbook()

    # ── "All Months" consolidated sheet ──────────────────────────
    ws_all = wb.active
    ws_all.title = "All Months"
    _write_header(ws_all, ALL_COLS, ALL_W)

    ri = 2
    for m in sorted(month_data.keys()):
        rows = month_data[m]
        if not rows:
            continue
        month_label = datetime(year, m, 1).strftime("%B %Y")
        band_hex    = MONTH_BAND_COLORS[(m - 1) % 12]
        band_fill   = PatternFill(start_color=band_hex, end_color=band_hex, fill_type="solid")
        for row in rows:
            row_with_label = dict(row, _month_label=month_label)
            _write_row(ws_all, ri, row_with_label, ALL_KEYS,
                       colour_map[row["ticketId"]], row_border(colour_map[row["ticketId"]]),
                       month_band_fill=band_fill)
            ri += 1

    if ri > 2:
        ws_all.auto_filter.ref = f"A1:{get_column_letter(len(ALL_COLS))}1"
    ws_all.freeze_panes = "A2"

    # ── One sheet per month ───────────────────────────────────────
    for m in sorted(month_data.keys()):
        rows = month_data[m]
        if not rows:
            continue
        ws_m = wb.create_sheet(title=datetime(year, m, 1).strftime("%b %Y"))
        _write_header(ws_m, COLS, COL_W)
        for ri2, row in enumerate(rows, 2):
            hex_c = colour_map[row["ticketId"]]
            _write_row(ws_m, ri2, row, COL_KEYS, hex_c, row_border(hex_c))
        ws_m.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
        ws_m.freeze_panes = "A2"

    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    with open(output_path, "wb") as f:
        f.write(buf.getvalue())
    buf.seek(0)
    return buf, filename


def main():
    try:
        check_env()
    except EnvironmentError as e:
        print(f"Error: {e}\nCreate a .env file — see .env.example")
        sys.exit(1)

    print("=== Jira Timesheet Generator ===")
    now = datetime.now()

    month_input = input(f"\nEnter month (1-12) [{now.month}]: ").strip()
    month = int(month_input) if month_input else now.month

    year_input = input(f"Enter year [{now.year}]: ").strip()
    year = int(year_input) if year_input else now.year

    if not 1 <= month <= 12:
        print("Invalid month. Must be between 1 and 12.")
        sys.exit(1)

    print(f"\nGenerating timesheet for {datetime(year, month, 1).strftime('%B %Y')}...")
    rows, errors = fetch_tickets(month, year)

    for err in errors:
        print(f"  Warning: {err}")

    if not rows:
        print("No tickets found for the selected month.")
        return

    _, filename = export_to_excel(rows, month, year)
    total_days = sum(r["duration"] for r in rows if isinstance(r["duration"], int))
    print(f"\n  Saved: {filename}")
    print(f"  Tickets: {len(rows)}  |  Total days: {total_days}")


if __name__ == "__main__":
    main()
