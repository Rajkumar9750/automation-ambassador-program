#!/usr/bin/env python3
"""
Dashboard Factory Tool
A modern web interface for modifying Tableau workbook formatting with real-time preview.
"""

import os
import sys
import logging

# Suppress Flask startup messages before importing Flask
logging.getLogger('werkzeug').setLevel(logging.ERROR)

from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import os
import sys
import json
import traceback
from werkzeug.utils import secure_filename
import threading
import uuid
from datetime import datetime
import shutil
import subprocess

# Add the current directory to the path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Add parent directory to path for accessing generate_filter_titles_report.py
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Import the formatter classes
try:
    from Filters import TableauFormatter
    from worksheet_title import WorksheetTitleModifier
    from change_filter_values_to_relevant import change_filter_values
    from change_filters_multiple_dropdown_with_apply import change_to_multiple_dropdown
    # Import dashboard size modifier and hide unused worksheets from parent directory
    import sys
    import os
    parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    sys.path.insert(0, parent_dir)
    from dashboard_size import DashboardSizeModifier
    from hide_unused_worksheets import UnusedWorksheetHider
    from delete_phone_layouts import PhoneLayoutDeleter
    sys.path.insert(0, parent_dir)
    from find_calculated_fields import CalculatedFieldsFinder
    try:
        from checks.dashboard_title_updater import DashboardTitleUpdater  # type: ignore
        DB_UPDATER_AVAILABLE = True
    except ImportError:
        DB_UPDATER_AVAILABLE = False
except Exception as e:
    DB_UPDATER_AVAILABLE = False

app = Flask(__name__)

# Disable Flask's default logger
app.logger.disabled = True
logging.getLogger('flask').setLevel(logging.ERROR)
logging.getLogger('werkzeug').setLevel(logging.ERROR)

# Enable CORS for all routes
CORS(app, 
     resources={r"/api/*": {"origins": "*", "methods": ["GET", "POST", "OPTIONS"]}},
     allow_headers=['Content-Type', 'Authorization'],
     support_credentials=True)

# Configuration
UPLOAD_FOLDER = '/tmp/tableau_uploads'
ALLOWED_EXTENSIONS = {'twb', 'twbx'}
MAX_FILE_SIZE = 500 * 1024 * 1024  # 500MB

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# Global state for tracking jobs
jobs = {}

# Add CORS headers to all responses
@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    response.headers.add('Access-Control-Max-Age', '3600')
    return response

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def validate_tableau_file(filepath):
    """Validate that the Tableau file is not corrupted."""
    try:
        import zipfile
        
        if filepath.lower().endswith('.twbx'):
            # Validate TWBX file
            with zipfile.ZipFile(filepath, 'r') as z:
                # Test all files in the archive
                result = z.testzip()
                if result is not None:
                    print(f"⚠️  Corrupted file in archive: {result}")
                    return False
                
                # Check that a .twb file exists
                twb_found = False
                for name in z.namelist():
                    if name.endswith('.twb'):
                        twb_found = True
                        break
                
                if not twb_found:
                    print(f"⚠️  No .twb file found in TWBX archive")
                    return False
        else:
            # For .twb files, just check file size
            if os.path.getsize(filepath) < 100:
                print(f"⚠️  File size too small: {os.path.getsize(filepath)} bytes")
                return False
        
        print(f"✅ File validation passed")
        return True
    except Exception as e:
        print(f"❌ File validation error: {e}")
        return False

@app.route('/')
def index():
    """Render the main page."""
    try:
        return render_template('index.html')
    except Exception as e:
        print(f"❌ Error rendering template: {e}")
        traceback.print_exc()
        return jsonify({'error': f'Template error: {str(e)}'}), 500

@app.route('/calculated-fields')
def calculated_fields():
    """Render the Calculated Fields tool page."""
    try:
        return render_template('calculated_fields.html')
    except Exception as e:
        print(f"❌ Error rendering calculated fields template: {e}")
        traceback.print_exc()
        return jsonify({'error': f'Template error: {str(e)}'}), 500

@app.route('/api/upload', methods=['POST', 'OPTIONS'])
def upload_file():
    """Handle file upload with browser compatibility."""
    
    # Handle preflight requests
    if request.method == 'OPTIONS':
        response = jsonify({'status': 'ok'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
        response.headers.add('Access-Control-Allow-Methods', 'GET,POST,OPTIONS')
        return response, 200
    
    try:
        print(f"📥 Upload request from {request.remote_addr}")
        print(f"   Content-Type: {request.content_type}")
        print(f"   Files in request: {list(request.files.keys())}")
        
        if 'file' not in request.files:
            print("❌ No file in request.files")
            return jsonify({'error': 'No file provided', 'status': 'error'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            print("❌ Empty filename")
            return jsonify({'error': 'No file selected', 'status': 'error'}), 400
        
        # Check file extension
        filename_lower = file.filename.lower()
        if not (filename_lower.endswith('.twb') or filename_lower.endswith('.twbx')):
            print(f"❌ Invalid file type: {file.filename}")
            return jsonify({'error': 'Only .twb and .twbx files allowed', 'status': 'error'}), 400
        
        # Generate unique filename
        unique_id = str(uuid.uuid4())[:8]
        filename = secure_filename(f"{unique_id}_{file.filename}")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        print(f"✅ Saving file to: {filepath}")
        
        # Save file
        file.save(filepath)
        
        # Verify file was saved
        if not os.path.exists(filepath):
            print(f"❌ File not saved: {filepath}")
            return jsonify({'error': 'Failed to save file', 'status': 'error'}), 500
        
        # Get file info
        file_size = os.path.getsize(filepath)
        print(f"✅ File saved successfully ({file_size} bytes)")
        
        response = jsonify({
            'success': True,
            'status': 'success',
            'file_id': unique_id,
            'filename': file.filename,
            'original_filename': file.filename,
            'filepath': filepath,
            'file_size': f"{file_size / (1024*1024):.2f} MB"
        })
        response.status_code = 200
        return response
        
    except Exception as e:
        print(f"❌ Exception during upload: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e), 'status': 'error'}), 500

@app.route('/api/format', methods=['POST'])
def format_workbook():
    """Apply formatting to workbook."""
    try:
        data = request.json
        filepath = data.get('filepath')
        config = data.get('config', {})
        
        print(f"📥 Format request - filepath: {filepath}")
        
        if not filepath or not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        # Create a unique job ID
        job_id = str(uuid.uuid4())[:8]
        jobs[job_id] = {'status': 'processing', 'progress': 0}
        
        # Process in background
        def process():
            try:
                print(f"📝 Formatting config: {list(config.keys())}")
                
                formatter = TableauFormatter(filepath)
                if not formatter.apply_formatting(config):
                    jobs[job_id] = {
                        'status': 'error',
                        'message': 'Failed to apply formatting'
                    }
                    return
                
                if not formatter.save_workbook():
                    jobs[job_id] = {
                        'status': 'error',
                        'message': 'Failed to save workbook'
                    }
                    return
                
                workbook_name = os.path.splitext(os.path.basename(filepath))[0]
                formatted_workbooks_path = os.path.expanduser("~/Desktop/Formatted_Workbooks")
                output_dir = os.path.join(formatted_workbooks_path, workbook_name)
                jobs[job_id] = {
                    'status': 'success',
                    'progress': 100,
                    'output_dir': output_dir,
                    'message': f'Success! File saved to: ~/Desktop/Formatted_Workbooks/{workbook_name}/'
                }
                print(f"✅ Formatting job {job_id} complete: {output_dir}")
            except Exception as e:
                print(f"❌ Error during formatting: {e}")
                traceback.print_exc()
                jobs[job_id] = {
                    'status': 'error',
                    'message': str(e),
                    'traceback': traceback.format_exc()
                }
        
        thread = threading.Thread(target=process)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'success': True,
            'job_id': job_id
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/format-worksheet-titles', methods=['POST'])
def format_worksheet_titles():
    """Apply worksheet title formatting to workbook."""
    try:
        data = request.json
        filepath = data.get('filepath')
        
        # Formatting options
        font_size = data.get('font_size', '15')
        font_family = data.get('font_family', 'Calibre')
        color = data.get('color', '#435254')
        font_style = data.get('font_style', 'normal')
        
        print(f"📥 Worksheet title format request - filepath: {filepath}")
        print(f"   Font Size: {font_size}, Font Family: {font_family}, Color: {color}, Style: {font_style}")
        
        if not filepath or not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        # Create a unique job ID
        job_id = str(uuid.uuid4())[:8]
        jobs[job_id] = {'status': 'processing', 'progress': 0}
        
        # Process in background
        def process():
            try:
                # Create output directory structure
                workbook_name = os.path.splitext(os.path.basename(filepath))[0]
                output_base_dir = os.path.dirname(filepath)
                output_dir = os.path.join(output_base_dir, workbook_name)
                os.makedirs(output_dir, exist_ok=True)
                
                # Prepare output file path
                file_extension = os.path.splitext(filepath)[1]
                output_filepath = os.path.join(output_dir, f"{workbook_name}{file_extension}")
                
                # Apply modifications and save to output path
                modifier = WorksheetTitleModifier(filepath)
                modified_count = modifier.apply_modifications(
                    output_path=output_filepath,
                    font_size=font_size,
                    font_family=font_family,
                    color=color,
                    font_style=font_style
                )
                
                jobs[job_id] = {
                    'status': 'success',
                    'progress': 100,
                    'output_dir': output_dir,
                    'output_file': output_filepath,
                    'elements_modified': modified_count,
                    'message': f'Success! {modified_count} elements modified. File saved to: {output_dir}/'
                }
                print(f"✅ Worksheet title formatting job {job_id} complete: {modified_count} elements modified")
                print(f"   Output saved to: {output_filepath}")
            except Exception as e:
                print(f"❌ Error during worksheet title formatting: {e}")
                traceback.print_exc()
                jobs[job_id] = {
                    'status': 'error',
                    'message': str(e),
                    'traceback': traceback.format_exc()
                }
        
        thread = threading.Thread(target=process)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'success': True,
            'job_id': job_id
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/process', methods=['POST'])
def process_workbook():
    """Apply filters and/or dashboard title formatting to workbook."""
    try:
        data = request.json
        filepath = data.get('filepath')
        original_filename = data.get('original_filename', '')  # Use original filename without UUID prefix
        apply_filters = data.get('apply_filters', False)
        apply_dashboard_titles = data.get('apply_dashboard_titles', False)
        apply_worksheet_titles = data.get('apply_worksheet_titles', False)
        apply_dashboard_sizes = data.get('apply_dashboard_sizes', False)
        apply_hide_used_worksheets = data.get('apply_hide_used_worksheets', False)
        apply_delete_phone_layouts = data.get('apply_delete_phone_layouts', False)
        apply_filter_values_relevant = data.get('apply_filter_values_relevant', False)
        apply_filter_type_dropdown = data.get('apply_filter_type_dropdown', False)
        filters_config = data.get('filters_config', {})
        format_configs = data.get('format_configs', {})  # Get custom format configs from frontend
        worksheet_titles_config = data.get('worksheet_titles_config', {
            'font_size': '15',
            'font_family': 'Calibre',
            'color': '#435254',
            'font_style': 'normal'
        })
        dashboard_config = data.get('dashboard_config', {
            'font_name': 'Calibre Medium',
            'font_size': '30',
            'font_color': '#012a2d'
        })
        dashboard_sizes_config = data.get('dashboard_sizes_config', {
            'width': '1366',
            'height': '1000',
            'sizing_mode': 'fixed'
        })
        
        # Ensure format_configs has default values for filters and parameters
        if 'filters' not in format_configs:
            format_configs['filters'] = {
                'titleColor': '#435254',
                'titleFont': 'Calibre',
                'titleSize': '11',
                'valueColor': '#000000',
                'valueFont': 'Calibre',
                'valueSize': '10'
            }
        
        if 'parameters' not in format_configs:
            format_configs['parameters'] = {
                'titleColor': '#435254',
                'titleFont': 'Calibre',
                'titleSize': '11',
                'valueColor': '#000000',
                'valueFont': 'Calibre',
                'valueSize': '10'
            }
        
        if 'dashboard-titles' not in format_configs:
            format_configs['dashboard-titles'] = {
                'font': 'Calibre Medium',
                'size': '30',
                'color': '#012a2d',
                'bold': True
            }
        
        if 'worksheet-titles' not in format_configs:
            format_configs['worksheet-titles'] = {
                'fontSize': '15',
                'fontFamily': 'Calibre',
                'color': '#435254',
                'fontStyle': 'normal'
            }
        
        # Override worksheet titles config if available
        if format_configs and 'worksheet-titles' in format_configs:
            worksheet_titles_config = {
                'font_size': str(format_configs['worksheet-titles'].get('fontSize', '15')),
                'font_family': format_configs['worksheet-titles'].get('fontFamily', 'Calibre'),
                'color': format_configs['worksheet-titles'].get('color', '#435254').upper() if format_configs['worksheet-titles'].get('color') else '#435254',
                'font_style': format_configs['worksheet-titles'].get('fontStyle', 'normal')
            }
        
        # Override dashboard_config with custom format_configs if available
        if format_configs and 'dashboard-titles' in format_configs:
            custom_dash = format_configs['dashboard-titles']
            dashboard_config['font_name'] = custom_dash.get('font', 'Calibre Medium')
            dashboard_config['font_size'] = str(custom_dash.get('size', '30'))
            # Normalize color to uppercase for Tableau XML
            color = custom_dash.get('color', '#012a2d')
            dashboard_config['font_color'] = color.upper() if color else '#012A2D'
        else:
            # Apply defaults if no custom config
            dashboard_config['font_name'] = 'Calibre Medium'
            dashboard_config['font_size'] = '30'
            dashboard_config['font_color'] = '#012A2D'
        
        print(f"   Dashboard Config: font_name={dashboard_config['font_name']}, font_size={dashboard_config['font_size']}, font_color={dashboard_config['font_color']}")
        # Override filter configs with custom format_configs if available
        if format_configs and 'filters' in format_configs:
            custom_filters = format_configs['filters']
            # Update filters_config with custom values
            filters_config.update({
                'quick-filter-title': {
                    'color': custom_filters.get('titleColor', '#435254'),
                    'font-family': custom_filters.get('titleFont', 'Calibre'),
                    'font-size': custom_filters.get('titleSize', '11')
                },
                'quick-filter': {
                    'color': custom_filters.get('valueColor', '#000000'),
                    'font-family': custom_filters.get('valueFont', 'Calibre'),
                    'font-size': custom_filters.get('valueSize', '10')
                }
            })
        
        # Override parameter configs with custom format_configs if available
        if format_configs and 'parameters' in format_configs:
            custom_params = format_configs['parameters']
            # Update filters_config with custom parameter values
            filters_config.update({
                'parameter-ctrl-title': {
                    'color': custom_params.get('titleColor', '#435254'),
                    'font-family': custom_params.get('titleFont', 'Calibre'),
                    'font-size': custom_params.get('titleSize', '11')
                },
                'parameter-ctrl': {
                    'color': custom_params.get('valueColor', '#000000'),
                    'font-family': custom_params.get('valueFont', 'Calibre'),
                    'font-size': custom_params.get('valueSize', '10')
                }
            })
        
        print(f"📥 Process request - filepath: {filepath}")
        print(f"   Apply Filters: {apply_filters}")
        print(f"   Apply Dashboard Titles: {apply_dashboard_titles}")
        print(f"   Apply Worksheet Titles: {apply_worksheet_titles}")
        print(f"   Apply Dashboard Sizes: {apply_dashboard_sizes}")
        print(f"   Apply Hide Used Worksheets: {apply_hide_used_worksheets}")
        print(f"   Apply Delete Phone Layouts: {apply_delete_phone_layouts}")
        print(f"   Apply Filter Values → Relevant: {apply_filter_values_relevant}")
        print(f"   Apply Filter Type → Dropdown: {apply_filter_type_dropdown}")
        print(f"   Custom Format Configs: {format_configs}")
        
        if not filepath or not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        if not apply_filters and not apply_dashboard_titles and not apply_worksheet_titles and not apply_dashboard_sizes and not apply_hide_used_worksheets and not apply_delete_phone_layouts and not apply_filter_values_relevant and not apply_filter_type_dropdown:
            return jsonify({'error': 'Please select at least one operation'}), 400
        
        # Create a unique job ID
        job_id = str(uuid.uuid4())[:8]
        jobs[job_id] = {'status': 'processing', 'progress': 0, 'steps': []}
        
        # Process in background
        def process():
            try:
                formatted_workbooks_path = os.path.expanduser("~/Desktop/Formatted_Workbooks")
                current_file = filepath
                output_dir = None
                actual_workbook_name = None
                
                print(f"\n{'='*80}")
                print(f"PROCESSING JOB {job_id}")
                print(f"{'='*80}")
                print(f"Input file: {filepath}")
                print(f"Apply Filters: {apply_filters}")
                print(f"Apply Dashboard Titles: {apply_dashboard_titles}")
                print(f"Apply Worksheet Titles: {apply_worksheet_titles}")
                
                # Step 1: Apply Filters
                if apply_filters:
                    jobs[job_id]['progress'] = 25
                    jobs[job_id]['steps'].append({'name': 'Applying Filters', 'status': 'in-progress'})
                    print(f"\n📝 STEP 1: Applying filters formatting...")
                    
                    try:
                        formatter = TableauFormatter(current_file)
                        print(f"   Formatter created, is_twbx: {formatter.is_twbx}")
                        print(f"   Filters config: {filters_config}")
                        
                        if not formatter.apply_formatting(filters_config):
                            jobs[job_id]['status'] = 'error'
                            jobs[job_id]['steps'][-1]['status'] = 'failed'
                            jobs[job_id]['message'] = 'Failed to apply filters'
                            print(f"❌ apply_formatting returned False")
                            return
                        
                        print(f"   Formatting applied successfully")
                        
                        if not formatter.save_workbook():
                            jobs[job_id]['status'] = 'error'
                            jobs[job_id]['steps'][-1]['status'] = 'failed'
                            jobs[job_id]['message'] = 'Failed to save after filters'
                            print(f"❌ save_workbook returned False")
                            return
                        
                        print(f"   Workbook saved successfully")
                    except Exception as e:
                        print(f"❌ Exception in filters step: {e}")
                        import traceback
                        traceback.print_exc()
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = f'Filter error: {str(e)}'
                        return
                    
                    # Get the actual workbook name from the original filename (without UUID prefix)
                    # But first, find where Filters.py actually put the output (it uses the uploaded filename with UUID)
                    uploaded_filename = os.path.basename(filepath)
                    uploaded_filename_no_ext = os.path.splitext(uploaded_filename)[0]
                    
                    # Filters.py creates a directory with the uploaded filename (including UUID prefix)
                    upload_output_dir = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_filename_no_ext)
                    output_twbx_path = os.path.join(upload_output_dir, f"{uploaded_filename_no_ext}.twbx")
                    
                    # Extract the clean workbook name (without UUID) for final output directory
                    if original_filename:
                        actual_workbook_name = os.path.splitext(original_filename)[0]
                    else:
                        # Remove UUID prefix if present (format: "uuid_originalname.ext")
                        if '_' in uploaded_filename_no_ext and len(uploaded_filename_no_ext.split('_')[0]) == 8:
                            actual_workbook_name = os.path.splitext('_'.join(uploaded_filename_no_ext.split('_')[1:]))[0]
                        else:
                            actual_workbook_name = uploaded_filename_no_ext
                    
                    print(f"\n   Looking for output at: {output_twbx_path}")
                    if not os.path.exists(output_twbx_path):
                        print(f"⚠️  Output file not found at {output_twbx_path}")
                        if os.path.exists(upload_output_dir):
                            print(f"   Directory contents: {os.listdir(upload_output_dir)}")
                        else:
                            print(f"   Directory does not exist: {upload_output_dir}")
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = f'Output file not created. Check logs.'
                        return
                    
                    # Copy to Formatted_Workbooks for final output
                    output_dir = os.path.join(formatted_workbooks_path, actual_workbook_name)
                    os.makedirs(output_dir, exist_ok=True)
                    final_output_path = os.path.join(output_dir, f"{actual_workbook_name}.twbx")
                    shutil.copy2(output_twbx_path, final_output_path)
                    
                    # Get the output file path
                    current_file = final_output_path
                    jobs[job_id]['steps'][-1]['status'] = 'completed'
                    print(f"✅ Filters applied successfully")
                    print(f"   Output: {final_output_path}")
                
                # Step 2: Change Filter Values to Relevant
                if apply_filter_values_relevant:
                    jobs[job_id]['progress'] = 40
                    jobs[job_id]['steps'].append({'name': 'Changing Filter Values to Relevant', 'status': 'in-progress'})
                    print(f"\n📝 STEP 2: Changing filter values to relevant...")
                    
                    try:
                        # Ensure output_dir and actual_workbook_name are set
                        if not output_dir or not actual_workbook_name:
                            # Need to initialize these if filters weren't applied
                            if original_filename:
                                actual_workbook_name = os.path.splitext(original_filename)[0]
                            else:
                                uploaded_filename = os.path.basename(filepath)
                                uploaded_filename_no_ext = os.path.splitext(uploaded_filename)[0]
                                if '_' in uploaded_filename_no_ext and len(uploaded_filename_no_ext.split('_')[0]) == 8:
                                    actual_workbook_name = os.path.splitext('_'.join(uploaded_filename_no_ext.split('_')[1:]))[0]
                                else:
                                    actual_workbook_name = uploaded_filename_no_ext
                            
                            output_dir = os.path.join(formatted_workbooks_path, actual_workbook_name)
                            os.makedirs(output_dir, exist_ok=True)
                            
                            # Copy the original file if this is the first operation
                            if not apply_filters:
                                current_file = os.path.join(output_dir, f"{actual_workbook_name}.twbx")
                                shutil.copy2(filepath, current_file)
                                print(f"   Copied original file to: {current_file}")
                        
                        # Call change_filter_values with the output_dir
                        change_filter_values(current_file, output_dir)
                        
                        # Update current_file to point to the output file
                        current_file = os.path.join(output_dir, os.path.basename(current_file))
                        print(f"✅ Filter values changed successfully")
                    except Exception as e:
                        print(f"❌ Exception in filter values step: {e}")
                        import traceback
                        traceback.print_exc()
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = f'Filter values error: {str(e)}'
                        return
                    
                    jobs[job_id]['steps'][-1]['status'] = 'completed'
                
                # Step 3: Change Filter Type to Dropdown with Apply
                if apply_filter_type_dropdown:
                    jobs[job_id]['progress'] = 60
                    jobs[job_id]['steps'].append({'name': 'Converting Filters to Multiple Dropdown with Apply', 'status': 'in-progress'})
                    print(f"\n📝 STEP 3: Converting filters to multiple dropdown with apply...")
                    
                    try:
                        # Ensure output_dir and actual_workbook_name are set
                        if not output_dir or not actual_workbook_name:
                            # Need to initialize these if filters weren't applied
                            if original_filename:
                                actual_workbook_name = os.path.splitext(original_filename)[0]
                            else:
                                uploaded_filename = os.path.basename(filepath)
                                uploaded_filename_no_ext = os.path.splitext(uploaded_filename)[0]
                                if '_' in uploaded_filename_no_ext and len(uploaded_filename_no_ext.split('_')[0]) == 8:
                                    actual_workbook_name = os.path.splitext('_'.join(uploaded_filename_no_ext.split('_')[1:]))[0]
                                else:
                                    actual_workbook_name = uploaded_filename_no_ext
                            
                            output_dir = os.path.join(formatted_workbooks_path, actual_workbook_name)
                            os.makedirs(output_dir, exist_ok=True)
                            
                            # Copy the original file if this is the first operation
                            if not apply_filters and not apply_filter_values_relevant:
                                current_file = os.path.join(output_dir, f"{actual_workbook_name}.twbx")
                                shutil.copy2(filepath, current_file)
                                print(f"   Copied original file to: {current_file}")
                        
                        # Call change_to_multiple_dropdown with the output_dir
                        change_to_multiple_dropdown(current_file, output_dir)
                        
                        # Update current_file to point to the output file
                        current_file = os.path.join(output_dir, os.path.basename(current_file))
                        print(f"✅ Filter types converted successfully")
                    except Exception as e:
                        print(f"❌ Exception in filter type conversion step: {e}")
                        import traceback
                        traceback.print_exc()
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = f'Filter type conversion error: {str(e)}'
                        return
                    
                    jobs[job_id]['steps'][-1]['status'] = 'completed'
                
                # Step 4: Apply Dashboard Titles
                if apply_dashboard_titles:
                    jobs[job_id]['progress'] = 80
                    jobs[job_id]['steps'].append({'name': 'Updating Dashboard Titles', 'status': 'in-progress'})
                    print(f"\n📝 STEP 4: Applying dashboard title formatting...")
                    
                    # If only dashboard titles (no filters), we need to extract the workbook name and prepare output
                    if not apply_filters:
                        # Extract temp directory to get the actual workbook name
                        import tempfile
                        import zipfile
                        temp_extract = tempfile.mkdtemp()
                        try:
                            with zipfile.ZipFile(filepath, 'r') as zip_ref:
                                zip_ref.extractall(temp_extract)
                            # Find the .twb file
                            for root, dirs, files in os.walk(temp_extract):
                                for file in files:
                                    if file.endswith('.twb'):
                                        actual_workbook_name = os.path.splitext(file)[0]
                                        output_dir = os.path.join(formatted_workbooks_path, actual_workbook_name)
                                        os.makedirs(output_dir, exist_ok=True)
                                        # Copy the ORIGINAL file to output directory (clean, no modifications)
                                        current_file = os.path.join(output_dir, f"{actual_workbook_name}.twbx")
                                        shutil.copy2(filepath, current_file)
                                        print(f"   Copied original file to: {current_file}")
                                        break
                        finally:
                            shutil.rmtree(temp_extract, ignore_errors=True)
                    
                    if not os.path.exists(current_file):
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = f'File not found: {current_file}'
                        print(f"❌ Input file not found: {current_file}")
                        return
                    
                    print(f"   Processing: {current_file}")
                    
                    try:
                        # Format dashboard titles using the dashboard_title.py script
                        font_name = dashboard_config.get('font_name', 'Calibre Medium')
                        font_size = dashboard_config.get('font_size', '30')
                        font_color = dashboard_config.get('font_color', '#012A2D')
                        # Ensure color is uppercase for Tableau
                        if font_color and font_color.startswith('#'):
                            font_color = font_color.upper()
                        
                        print(f"   Dashboard Title Formatting Parameters:")
                        print(f"      Font: {font_name}")
                        print(f"      Size: {font_size}pt")
                        print(f"      Color: {font_color}")
                        
                        # Call the dashboard_title.py script
                        script_dir = os.path.dirname(os.path.abspath(__file__))
                        dashboard_script = os.path.join(script_dir, 'dashboard_title.py')
                        
                        cmd = [
                            sys.executable,
                            dashboard_script,
                            current_file,
                            font_name,
                            font_size,
                            font_color
                        ]
                        print(f"   Running command: {' '.join(cmd)}")
                        
                        result = subprocess.run(cmd, capture_output=True, text=True, cwd=script_dir)
                        
                        if result.stderr:
                            print(f"   Debug output:\n{result.stderr}")
                        
                        success = result.returncode == 0
                        
                        if not success:
                            jobs[job_id]['status'] = 'error'
                            jobs[job_id]['steps'][-1]['status'] = 'failed'
                            jobs[job_id]['message'] = 'Failed to update dashboard titles'
                            print(f"❌ Dashboard title formatting script failed: {result.stderr}")
                            return
                        
                        print(f"✅ Dashboard titles formatted successfully")
                        jobs[job_id]['steps'][-1]['status'] = 'completed'
                        jobs[job_id]['progress'] = 90
                    except Exception as e:
                        print(f"❌ Exception in dashboard titles step: {e}")
                        import traceback
                        traceback.print_exc()
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = f'Dashboard title error: {str(e)}'
                        return
                    
                    jobs[job_id]['steps'][-1]['status'] = 'completed'
                    jobs[job_id]['progress'] = 100
                
                # Step 5: Apply Worksheet Titles Formatting
                if apply_worksheet_titles:
                    jobs[job_id]['progress'] = 85
                    jobs[job_id]['steps'].append({'name': 'Formatting Worksheet Titles', 'status': 'in-progress'})
                    print(f"\n📝 STEP 5: Applying worksheet title formatting...")
                    
                    if not os.path.exists(current_file):
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = f'File not found for worksheet titles: {current_file}'
                        print(f"❌ Input file not found: {current_file}")
                        return
                    
                    try:
                        print(f"   Processing: {current_file}")
                        print(f"   Worksheet Title Formatting Parameters:")
                        print(f"      Font Size: {worksheet_titles_config.get('font_size', '15')}pt")
                        print(f"      Font Family: {worksheet_titles_config.get('font_family', 'Calibre')}")
                        print(f"      Color: {worksheet_titles_config.get('color', '#435254')}")
                        print(f"      Font Style: {worksheet_titles_config.get('font_style', 'normal')}")
                        
                        modifier = WorksheetTitleModifier(current_file)
                        modified_count = modifier.apply_modifications(
                            output_path=current_file,
                            font_size=worksheet_titles_config.get('font_size', '15'),
                            font_family=worksheet_titles_config.get('font_family', 'Calibre'),
                            color=worksheet_titles_config.get('color', '#435254'),
                            font_style=worksheet_titles_config.get('font_style', 'normal')
                        )
                        
                        print(f"✅ Worksheet titles formatted successfully ({modified_count} elements modified)")
                        jobs[job_id]['steps'][-1]['status'] = 'completed'
                        jobs[job_id]['progress'] = 95
                    except Exception as e:
                        print(f"❌ Exception in worksheet titles step: {e}")
                        import traceback
                        traceback.print_exc()
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = f'Worksheet title error: {str(e)}'
                        return
                
                if apply_dashboard_sizes:
                    jobs[job_id]['progress'] = 90
                    jobs[job_id]['steps'].append({'name': 'Modifying Dashboard Sizes', 'status': 'in-progress'})
                    print(f"\n📐 STEP 6: Modifying dashboard sizes...")
                    
                    if not os.path.exists(current_file):
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = f'File not found for dashboard sizes: {current_file}'
                        print(f"❌ Input file not found: {current_file}")
                        return
                    
                    try:
                        print(f"   Processing: {current_file}")
                        print(f"   Dashboard Size Parameters:")
                        print(f"      Width: {dashboard_sizes_config.get('width', '1366')}px")
                        print(f"      Height: {dashboard_sizes_config.get('height', '1000')}px")
                        print(f"      Sizing Mode: {dashboard_sizes_config.get('sizing_mode', 'fixed')}")
                        
                        modifier = DashboardSizeModifier(current_file)
                        modified_count, _ = modifier.apply_modifications(dashboard_sizes_config)
                        modifier.save_workbook(current_file)
                        
                        print(f"✅ Dashboard sizes modified successfully ({modified_count} dashboards updated)")
                        jobs[job_id]['steps'][-1]['status'] = 'completed'
                        jobs[job_id]['progress'] = 97
                    except Exception as e:
                        print(f"❌ Exception in dashboard sizes step: {e}")
                        import traceback
                        traceback.print_exc()
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = f'Dashboard size error: {str(e)}'
                        return
                
                # Step 7: Hide Used Worksheets
                if apply_hide_used_worksheets:
                    jobs[job_id]['progress'] = 93
                    jobs[job_id]['steps'].append({'name': 'Hiding Used Worksheets', 'status': 'in-progress'})
                    print(f"\n🙈 STEP 7: Hiding used worksheets...")
                    
                    if not os.path.exists(current_file):
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = f'File not found for hide worksheets: {current_file}'
                        print(f"❌ Input file not found: {current_file}")
                        return
                    
                    try:
                        print(f"   Processing: {current_file}")
                        hider = UnusedWorksheetHider(current_file)
                        hider.apply_modifications(current_file)
                        
                        print(f"✅ Used worksheets hidden successfully")
                        jobs[job_id]['steps'][-1]['status'] = 'completed'
                        jobs[job_id]['progress'] = 97
                    except Exception as e:
                        print(f"❌ Exception in hide worksheets step: {e}")
                        import traceback
                        traceback.print_exc()
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = f'Hide worksheets error: {str(e)}'
                        return
                
                # Step 8: Delete Phone Layouts
                if apply_delete_phone_layouts:
                    jobs[job_id]['progress'] = 97
                    jobs[job_id]['steps'].append({'name': 'Deleting Phone Layouts', 'status': 'in-progress'})
                    print(f"\n📱 STEP 8: Deleting phone layouts from dashboards...")
                    
                    if not os.path.exists(current_file):
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = f'File not found for delete phone layouts: {current_file}'
                        print(f"❌ Input file not found: {current_file}")
                        return
                    
                    try:
                        print(f"   Processing: {current_file}")
                        deleter = PhoneLayoutDeleter(current_file)
                        deleter.apply_modifications(current_file)
                        
                        print(f"✅ Phone layouts deleted successfully")
                        jobs[job_id]['steps'][-1]['status'] = 'completed'
                        jobs[job_id]['progress'] = 99
                    except Exception as e:
                        print(f"❌ Exception in delete phone layouts step: {e}")
                        import traceback
                        traceback.print_exc()
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = f'Delete phone layouts error: {str(e)}'
                        return
                
                
                # Success
                success_msg = f'✅ File saved to: {output_dir}'
                
                jobs[job_id] = {
                    'status': 'success',
                    'progress': 100,
                    'output_dir': output_dir,
                    'output_file': current_file,
                    'steps': jobs[job_id].get('steps', []),
                    'message': success_msg
                }
                
                # Validate the final workbook before declaring success
                print(f"\n📋 Final validation of output file...")
                if not validate_tableau_file(current_file):
                    print(f"⚠️  Warning: Output file validation failed")
                    jobs[job_id]['warning'] = 'Output file may be corrupted. Please verify in Tableau.'
                else:
                    print(f"✅ Output file validation passed")
                
                print(f"\n{'='*80}")
                print(f"✅ Processing job {job_id} COMPLETE")
                print(f"{'='*80}")
                print(f"Output directory: {output_dir}")
                print(f"Output file: {current_file}")
            except Exception as e:
                print(f"❌ Error during processing: {e}")
                traceback.print_exc()
                jobs[job_id] = {
                    'status': 'error',
                    'message': str(e),
                    'traceback': traceback.format_exc(),
                    'steps': jobs[job_id].get('steps', [])
                }
        
        thread = threading.Thread(target=process)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'success': True,
            'job_id': job_id
        }), 200
        
    except Exception as e:
        print(f"❌ Error in process_workbook: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/job-status/<job_id>', methods=['GET'])
def job_status(job_id):
    """Get the status of a formatting job."""
    if job_id not in jobs:
        return jsonify({'error': 'Job not found'}), 404
    
    return jsonify(jobs[job_id]), 200

@app.route('/api/config', methods=['GET'])
def get_config():
    """Get the CBRE configuration."""
    from Filters import get_default_cbre_config
    filters_config = get_default_cbre_config()
    
    # Dashboard titles are handled separately by update_dashboard_titles.py
    # Not included in Filters configuration
    return jsonify(filters_config), 200

@app.route('/api/get-calculated-fields', methods=['POST'])
def get_calculated_fields():
    """Get list of all calculated fields from uploaded workbook."""
    try:
        from find_calculated_fields import CalculatedFieldsFinder
        
        data = request.json
        filepath = data.get('filepath')
        
        if not filepath or not os.path.exists(filepath):
            return jsonify({'error': 'File not found', 'fields': []}), 404
        
        finder = CalculatedFieldsFinder(filepath)
        
        if not finder.extract_workbook():
            return jsonify({'error': 'Failed to extract workbook', 'fields': []}), 500
        
        if not finder.extract_calculated_fields():
            finder.cleanup()
            return jsonify({'error': 'Failed to extract calculated fields', 'fields': []}), 500
        
        # Build field list
        fields = []
        for field in finder.calculated_fields:
            fields.append({
                'name': field['name'],
                'caption': field['caption'],
                'datasource': field['datasource']
            })
        
        finder.cleanup()
        
        return jsonify({
            'success': True,
            'total_fields': len(fields),
            'fields': fields
        }), 200
        
    except Exception as e:
        print(f"Error getting calculated fields: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'fields': []}), 500

@app.route('/api/search-calculated-fields', methods=['POST'])
def search_calculated_fields():
    """Search for a term in all calculated fields of the workbook."""
    try:
        data = request.json
        filepath = data.get('filepath')
        search_term = data.get('search_term', '')
        
        if not filepath or not os.path.exists(filepath):
            return jsonify({'error': 'File not found', 'results': []}), 404
        
        if not search_term or len(search_term.strip()) == 0:
            return jsonify({'error': 'Search term cannot be empty', 'results': []}), 400
        
        print(f"\n🔍 Searching for calculated fields containing: '{search_term}'")
        
        try:
            finder = CalculatedFieldsFinder(filepath)
            results = finder.search(search_term)
            
            print(f"✅ Search complete: Found {results.get('matches_found', 0)} matching calculated fields ({results.get('exact_matches', 0)} exact)")
            
            return jsonify({
                'success': True,
                'search_term': search_term,
                'total_calculated_fields': results.get('total_calculated_fields', 0),
                'matches_found': results.get('matches_found', 0),
                'exact_matches': results.get('exact_matches', 0),
                'results': results.get('results', []),
                'error': results.get('error')
            }), 200
            
        except Exception as e:
            print(f"❌ Error searching calculated fields: {e}")
            traceback.print_exc()
            return jsonify({
                'success': False,
                'error': str(e),
                'results': []
            }), 500
            
    except Exception as e:
        print(f"❌ Error in search_calculated_fields endpoint: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e), 'results': []}), 500

@app.route('/api/download-calculated-fields', methods=['POST'])
def download_calculated_fields():
    """Generate and download Excel file with calculated fields search results."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        data = request.json
        search_term = data.get('search_term', 'Search Results')
        results = data.get('results', [])
        
        if not results:
            return jsonify({'error': 'No results to download'}), 400
        
        print(f"📊 Generating Excel file for {len(results)} calculated field(s)...")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Calculated Fields"
        
        # Define styles
        header_fill = PatternFill(start_color="00897B", end_color="00897B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        alt_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Add summary
        ws['A1'] = f"Calculated Fields Search Results"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A2'] = f"Search Term: {search_term}"
        ws['A2'].font = Font(italic=True, size=10)
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(italic=True, size=10)
        
        # Add header row (starting at row 5)
        headers = ['Field Name', 'Data Source', 'Calculation', 'Match In']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Add data rows
        for row_idx, result in enumerate(results, 6):
            # Field Name
            field_name = result.get('caption') or result.get('name', 'N/A')
            internal_name = result.get('name', '')
            display_name = f"{field_name}\n({internal_name})" if internal_name != field_name else field_name
            
            cell = ws.cell(row=row_idx, column=1)
            cell.value = display_name
            cell.alignment = left_align
            cell.border = border
            if (row_idx - 6) % 2 == 1:
                cell.fill = alt_fill
            
            # Data Source
            cell = ws.cell(row=row_idx, column=2)
            cell.value = result.get('datasource', 'N/A')
            cell.alignment = left_align
            cell.border = border
            if (row_idx - 6) % 2 == 1:
                cell.fill = alt_fill
            
            # Calculation (resolved formula)
            formula = result.get('resolved_formula') or result.get('formula', 'N/A')
            cell = ws.cell(row=row_idx, column=3)
            cell.value = formula
            cell.alignment = left_align
            cell.border = border
            if (row_idx - 6) % 2 == 1:
                cell.fill = alt_fill
            
            # Match In
            match_locations = []
            if result.get('match_location', {}).get('name'):
                match_locations.append('Name')
            if result.get('match_location', {}).get('caption'):
                match_locations.append('Caption')
            if result.get('match_location', {}).get('formula'):
                match_locations.append('Formula')
            
            cell = ws.cell(row=row_idx, column=4)
            cell.value = ', '.join(match_locations)
            cell.alignment = center_align
            cell.border = border
            if (row_idx - 6) % 2 == 1:
                cell.fill = alt_fill
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 20
        
        # Set row heights
        ws.row_dimensions[5].height = 20
        for row in range(6, 6 + len(results)):
            ws.row_dimensions[row].height = 30
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        print(f"✅ Excel file generated successfully with {len(results)} row(s)")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Calculated_Fields_{search_term}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        print(f"❌ Error generating Excel file: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-all-calculated-fields', methods=['POST'])
def download_all_calculated_fields():
    """Generate and download Excel file with ALL calculated fields from the workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        data = request.json
        filepath = data.get('filepath')
        
        if not filepath or not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        print(f"📊 Exporting ALL calculated fields to Excel...")
        
        # Extract all calculated fields
        try:
            finder = CalculatedFieldsFinder(filepath)
            if not finder.extract_workbook():
                return jsonify({'error': 'Failed to extract workbook'}), 400
            
            if not finder.extract_calculated_fields():
                return jsonify({'error': 'Failed to extract calculated fields'}), 400
            
            all_fields = finder.calculated_fields
            total_fields = len(all_fields)
            
            print(f"✓ Found {total_fields} calculated field(s)")
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "All Calculated Fields"
            
            # Define styles
            header_fill = PatternFill(start_color="1976d2", end_color="1976d2", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            alt_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Add summary
            ws['A1'] = f"All Calculated Fields - {os.path.basename(filepath)}"
            ws['A1'].font = Font(bold=True, size=12)
            ws['A2'] = f"Total Fields: {total_fields}"
            ws['A2'].font = Font(italic=True, size=10)
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'].font = Font(italic=True, size=10)
            
            # Add header row (starting at row 5)
            headers = ['Field Name', 'Display Name', 'Data Source', 'Calculation']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            
            # Add data rows
            for row_idx, field in enumerate(all_fields, 6):
                # Field Name (internal name)
                cell = ws.cell(row=row_idx, column=1)
                cell.value = field.get('name', 'N/A')
                cell.alignment = left_align
                cell.border = border
                if (row_idx - 6) % 2 == 1:
                    cell.fill = alt_fill
                
                # Display Name (caption)
                cell = ws.cell(row=row_idx, column=2)
                cell.value = field.get('caption', field.get('name', 'N/A'))
                cell.alignment = left_align
                cell.border = border
                if (row_idx - 6) % 2 == 1:
                    cell.fill = alt_fill
                
                # Data Source
                cell = ws.cell(row=row_idx, column=3)
                cell.value = field.get('datasource', 'N/A')
                cell.alignment = left_align
                cell.border = border
                if (row_idx - 6) % 2 == 1:
                    cell.fill = alt_fill
                
                # Calculation (resolved formula)
                formula = field.get('formula', 'N/A')
                # Resolve the formula to show readable field names
                resolved = finder.resolve_formula(formula)
                
                cell = ws.cell(row=row_idx, column=4)
                cell.value = resolved
                cell.alignment = left_align
                cell.border = border
                if (row_idx - 6) % 2 == 1:
                    cell.fill = alt_fill
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 50
            
            # Set row heights
            ws.row_dimensions[5].height = 20
            for row in range(6, 6 + len(all_fields)):
                ws.row_dimensions[row].height = 30
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            print(f"✅ Excel file generated successfully with {total_fields} calculated fields")
            
            # Cleanup
            finder.cleanup()
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'All_Calculated_Fields_{datetime.now().strftime("%Y%m%d")}.xlsx'
            )
            
        except Exception as e:
            print(f"❌ Error processing calculated fields: {e}")
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500
            
    except Exception as e:
        print(f"❌ Error in download_all_calculated_fields endpoint: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/presets', methods=['GET'])
def get_presets():
    """Get available formatting presets."""
    presets = {
        'default': {
            'name': 'Default',
            'description': 'Standard Tableau styling',
            'config': {
                'quick-filter-title': {
                    'color': '#11e033',
                    'font-family': 'Calibre',
                    'font-size': '12',
                    'font-weight': 'bold',
                    'background-color': '#f8f9fa'
                },
                'quick-filter': {
                    'color': '#D60404',
                    'font-family': 'Calibre',
                    'font-size': '10',
                    'background-color': '#ffffff',
                    'border-color': '#cad1d3'
                },
                'parameter-ctrl-title': {
                    'color': '#ed09d6',
                    'font-family': 'Calibre',
                    'font-size': '12',
                    'font-weight': 'bold',
                    'background-color': '#f8f9fa'
                },
                'parameter-ctrl': {
                    'color': '#280BE4',
                    'font-family': 'Calibre',
                    'font-size': '10',
                    'background-color': '#ffffff',
                    'border-color': '#cad1d3'
                }
            }
        },
        'professional': {
            'name': 'Professional Blue',
            'description': 'Professional blue color scheme',
            'config': {
                'quick-filter-title': {
                    'color': '#003f7f',
                    'font-weight': 'bold',
                    'font-size': '11'
                },
                'quick-filter': {
                    'color': '#0066cc',
                    'font-size': '10'
                },
                'parameter-ctrl-title': {
                    'color': '#003f7f',
                    'font-weight': 'bold',
                    'font-size': '11'
                },
                'parameter-ctrl': {
                    'color': '#0066cc',
                    'font-size': '10'
                }
            }
        },
        'vibrant': {
            'name': 'Vibrant',
            'description': 'Vibrant and bold colors',
            'config': {
                'quick-filter-title': {
                    'color': '#FF6B35',
                    'font-weight': 'bold',
                    'font-size': '12'
                },
                'quick-filter': {
                    'color': '#004E89',
                    'font-size': '10'
                },
                'parameter-ctrl-title': {
                    'color': '#1B998B',
                    'font-weight': 'bold',
                    'font-size': '12'
                },
                'parameter-ctrl': {
                    'color': '#C1121F',
                    'font-size': '10'
                }
            }
        },
        'minimal': {
            'name': 'Minimal',
            'description': 'Clean and minimal design',
            'config': {
                'quick-filter-title': {
                    'color': '#2c3e50',
                    'font-weight': 'bold',
                    'font-size': '11'
                },
                'quick-filter': {
                    'color': '#34495e',
                    'font-size': '10'
                },
                'parameter-ctrl-title': {
                    'color': '#2c3e50',
                    'font-weight': 'bold',
                    'font-size': '11'
                },
                'parameter-ctrl': {
                    'color': '#34495e',
                    'font-size': '10'
                }
            }
        }
    }
    
    # Add worksheet title presets
    worksheet_title_presets = {
        'worksheet_titles': {
            'title': 'Worksheet Title Formatting',
            'presets': {
                'default': {
                    'name': 'Default',
                    'description': 'Standard worksheet title styling',
                    'config': {
                        'font_size': '15',
                        'font_family': 'Calibre',
                        'color': '#435254',
                        'font_style': 'normal'
                    }
                },
                'bold': {
                    'name': 'Bold',
                    'description': 'Bold worksheet titles',
                    'config': {
                        'font_size': '15',
                        'font_family': 'Calibre',
                        'color': '#333333',
                        'font_style': 'bold'
                    }
                },
                'large': {
                    'name': 'Large',
                    'description': 'Large worksheet titles',
                    'config': {
                        'font_size': '18',
                        'font_family': 'Calibre',
                        'color': '#2c3e50',
                        'font_style': 'bold'
                    }
                },
                'stylized': {
                    'name': 'Stylized',
                    'description': 'Stylized with pink color',
                    'config': {
                        'font_size': '10',
                        'font_family': 'Calibre',
                        'color': '#FF69B4',
                        'font_style': 'normal'
                    }
                },
                'professional': {
                    'name': 'Professional',
                    'description': 'Professional blue titles',
                    'config': {
                        'font_size': '14',
                        'font_family': 'Calibre',
                        'color': '#003f7f',
                        'font_style': 'bold'
                    }
                }
            }
        }
    }
    
    # Combine both presets
    all_presets = {
        'filters': presets,
        'worksheet_titles': worksheet_title_presets['worksheet_titles']
    }
    
    return jsonify(all_presets), 200

@app.route('/api/elements', methods=['GET'])
def get_elements():
    """Get available Tableau elements for formatting."""
    elements = {
        'filters': {
            'name': 'Filters',
            'description': 'Filter controls (title and dropdown)',
            'subelements': ['quick-filter-title', 'quick-filter'],
            'example': 'Region, Product, Date'
        },
        'parameters': {
            'name': 'Parameters',
            'description': 'Parameter controls (title and dropdown)',
            'subelements': ['parameter-ctrl-title', 'parameter-ctrl'],
            'example': 'Select Year, Choose Option'
        },
        'worksheet_titles': {
            'name': 'Worksheet Titles',
            'description': 'Worksheet title formatting across all worksheets',
            'attributes': ['font_size', 'font_family', 'color', 'font_style'],
            'example': 'All worksheet titles will be formatted uniformly'
        }
    }
    
    return jsonify(elements), 200

@app.route('/api/attributes', methods=['GET'])
def get_attributes():
    """Get available formatting attributes."""
    attributes = {
        'color': {
            'name': 'Text Color',
            'type': 'color',
            'description': 'Color of the text',
            'default': '#000000'
        },
        'background-color': {
            'name': 'Background Color',
            'type': 'color',
            'description': 'Background color of the element',
            'default': '#ffffff'
        },
        'border-color': {
            'name': 'Border Color',
            'type': 'color',
            'description': 'Border color of the element',
            'default': '#cccccc'
        },
        'font-size': {
            'name': 'Font Size',
            'type': 'number',
            'min': 8,
            'max': 32,
            'description': 'Size of the font in points',
            'default': '10'
        },
        'font-family': {
            'name': 'Font Family',
            'type': 'select',
            'options': ['Arial', 'Calibre', 'Courier New', 'Times New Roman', 'Verdana', 'Georgia'],
            'description': 'Font typeface',
            'default': 'Calibre'
        },
        'font-weight': {
            'name': 'Font Weight',
            'type': 'select',
            'options': ['normal', 'bold'],
            'description': 'Font weight (normal or bold)',
            'default': 'normal'
        },
        'font_style': {
            'name': 'Font Style',
            'type': 'select',
            'options': ['normal', 'bold', 'italic'],
            'description': 'Font style (normal, bold, or italic)',
            'default': 'normal'
        }
    }
    
    return jsonify(attributes), 200

@app.route('/api/update-dashboard-formatting', methods=['POST'])
def update_dashboard_formatting():
    """
    Apply dashboard title formatting to uploaded workbook.
    Uses the DashboardTitleUpdater settings (Calibre, size 15, yellow).
    Saves to Desktop/Formatted_Workbooks folder.
    """
    try:
        data = request.json
        filepath = data.get('filepath')
        
        print(f"📥 Dashboard formatting request - filepath: {filepath}")
        
        if not filepath or not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        # Create a unique job ID
        job_id = str(uuid.uuid4())[:8]
        jobs[job_id] = {'status': 'processing', 'progress': 0, 'steps': []}
        
        # Process in background
        def process():
            try:
                print(f"\n{'='*80}")
                print(f"DASHBOARD FORMATTING JOB {job_id}")
                print(f"{'='*80}")
                
                jobs[job_id]['progress'] = 25
                jobs[job_id]['steps'].append({'name': 'Preparing workbook', 'status': 'in-progress'})
                
                # Get the actual workbook name from the uploaded filename
                import tempfile
                import zipfile
                
                temp_extract = tempfile.mkdtemp()
                actual_workbook_name = None
                
                try:
                    with zipfile.ZipFile(filepath, 'r') as zip_ref:
                        zip_ref.extractall(temp_extract)
                    # Find the .twb file
                    for root, dirs, files in os.walk(temp_extract):
                        for file in files:
                            if file.endswith('.twb'):
                                actual_workbook_name = os.path.splitext(file)[0]
                                break
                finally:
                    shutil.rmtree(temp_extract, ignore_errors=True)
                
                if not actual_workbook_name:
                    actual_workbook_name = os.path.splitext(os.path.basename(filepath))[0]
                
                # Create output directory
                formatted_workbooks_path = os.path.expanduser("~/Desktop/Formatted_Workbooks")
                output_dir = os.path.join(formatted_workbooks_path, actual_workbook_name)
                os.makedirs(output_dir, exist_ok=True)
                
                # Copy the file to output directory
                output_file = os.path.join(output_dir, f"{actual_workbook_name}.twbx")
                shutil.copy2(filepath, output_file)
                
                print(f"   Workbook copied to: {output_file}")
                jobs[job_id]['steps'][-1]['status'] = 'completed'
                
                # Apply dashboard title formatting
                jobs[job_id]['progress'] = 50
                jobs[job_id]['steps'].append({'name': 'Applying dashboard title formatting', 'status': 'in-progress'})
                
                print(f"\n📝 Applying dashboard title formatting...")
                print(f"   Font: Calibre | Size: 15 | Color: #00AA00 (Green)")
                
                try:
                    # Apply formatting using database-defined settings
                    result = subprocess.run([
                        'python3',
                        'format_all_titles_complete.py',
                        output_file,
                        'Calibre',
                        '15',
                        '#00AA00'
                    ], capture_output=True, text=True, cwd=os.path.dirname(os.path.abspath(__file__)))
                    
                    success = result.returncode == 0
                    
                    if not success:
                        jobs[job_id]['status'] = 'error'
                        jobs[job_id]['steps'][-1]['status'] = 'failed'
                        jobs[job_id]['message'] = 'Failed to apply dashboard title formatting'
                        print(f"❌ Dashboard title formatting script failed: {result.stderr}")
                        return
                    
                    # Find the generated file with _TITLES_ suffix and replace the original
                    output_dir_path = os.path.dirname(output_file)
                    workbook_base = os.path.splitext(os.path.basename(output_file))[0]
                    
                    # Find the generated file with _TITLES_ suffix
                    generated_files = [f for f in os.listdir(output_dir_path) if '_TITLES_' in f and f.endswith('.twbx')]
                    if generated_files:
                        # Use the most recently created file
                        generated_file = os.path.join(output_dir_path, max(generated_files, key=lambda f: os.path.getctime(os.path.join(output_dir_path, f))))
                        # Replace the original with the formatted version
                        os.remove(output_file)
                        shutil.move(generated_file, output_file)
                    
                    print(f"✅ Dashboard title formatting applied successfully")
                except Exception as e:
                    print(f"❌ Exception in dashboard formatting: {e}")
                    import traceback
                    traceback.print_exc()
                    jobs[job_id]['status'] = 'error'
                    jobs[job_id]['steps'][-1]['status'] = 'failed'
                    jobs[job_id]['message'] = f'Formatting error: {str(e)}'
                    return
                
                jobs[job_id]['steps'][-1]['status'] = 'completed'
                jobs[job_id]['progress'] = 100
                
                # Success
                jobs[job_id] = {
                    'status': 'success',
                    'progress': 100,
                    'output_dir': output_dir,
                    'output_file': output_file,
                    'workbook_name': actual_workbook_name,
                    'steps': jobs[job_id]['steps'],
                    'message': f'Success! Formatted workbook saved to:\n{output_dir}'
                }
                
                print(f"\n{'='*80}")
                print(f"✅ Formatting job {job_id} COMPLETE")
                print(f"{'='*80}")
                print(f"Output directory: {output_dir}")
                print(f"Output file: {output_file}")
                
            except Exception as e:
                print(f"❌ Error during formatting: {e}")
                traceback.print_exc()
                jobs[job_id] = {
                    'status': 'error',
                    'message': str(e),
                    'traceback': traceback.format_exc(),
                    'steps': jobs[job_id].get('steps', [])
                }
        
        thread = threading.Thread(target=process)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'success': True,
            'job_id': job_id
        }), 200
        
    except Exception as e:
        print(f"❌ Error in update_dashboard_formatting: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/db-update-dashboard-titles', methods=['POST'])
def db_update_dashboard_titles():
    """
    Update dashboard titles in database using structured query approach.
    Requires database connection and checks/dashboard_title_updater.py module.
    """
    if not DB_UPDATER_AVAILABLE:
        return jsonify({
            'error': 'Database updater module not available',
            'message': 'Please ensure checks/dashboard_title_updater.py is properly installed'
        }), 503
    
    try:
        data = request.json or {}
        
        # Create a unique job ID
        job_id = str(uuid.uuid4())[:8]
        jobs[job_id] = {'status': 'processing', 'progress': 0, 'updates': []}
        
        # Process in background
        def process():
            try:
                print(f"\n{'='*80}")
                print(f"DATABASE UPDATE JOB {job_id}")
                print(f"{'='*80}")
                
                # Import and initialize the updater
                # Note: This assumes db_manager is available in the environment
                # If not, the updater will handle the error appropriately
                updater = DashboardTitleUpdater()
                
                jobs[job_id]['progress'] = 50
                jobs[job_id]['steps'] = [{'name': 'Executing database updates', 'status': 'in-progress'}]
                
                print(f"\n📝 Executing dashboard title updates...")
                
                # Run the updates
                results = updater.run_update()
                
                jobs[job_id]['progress'] = 75
                
                # Process results
                update_count = len(results) if results else 0
                
                print(f"\n✅ Update job complete")
                print(f"   Total updates: {update_count}")
                
                if results:
                    for result in results:
                        print(f"   - {result.dashboard_name}: {result.attribute_updated} = {result.new_value}")
                
                jobs[job_id] = {
                    'status': 'success',
                    'progress': 100,
                    'steps': [{'name': 'Executing database updates', 'status': 'completed'}],
                    'message': f'Successfully updated {update_count} dashboard formatting rules',
                    'updates': [
                        {
                            'dashboard_id': r.dashboard_id,
                            'dashboard_name': r.dashboard_name,
                            'attribute': r.attribute_updated,
                            'old_value': r.old_value,
                            'new_value': r.new_value,
                            'status': r.status
                        } for r in (results or [])
                    ]
                }
                
                print(f"{'='*80}")
                
            except Exception as e:
                print(f"❌ Error during database update: {e}")
                traceback.print_exc()
                jobs[job_id] = {
                    'status': 'error',
                    'message': str(e),
                    'error_details': traceback.format_exc(),
                    'steps': jobs[job_id].get('steps', [])
                }
        
        thread = threading.Thread(target=process)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'success': True,
            'job_id': job_id,
            'message': 'Database update job started'
        }), 200
        
    except Exception as e:
        print(f"❌ Error in db_update_dashboard_titles: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint."""
    return jsonify({
        'status': 'ok',
        'db_updater_available': DB_UPDATER_AVAILABLE
    }), 200

@app.route('/api/format-dashboard-titles', methods=['POST'])
def format_dashboard_titles():
    """Format dashboard titles in Tableau workbook."""
    try:
        data = request.json
        filepath = data.get('filepath')
        original_filename = data.get('original_filename', '')
        fontname = data.get('fontname', 'Calibre Medium')
        fontsize = data.get('fontsize', '30')
        fontcolor = data.get('fontcolor', '#012A2D')
        bold = data.get('bold', True)
        
        print(f"\n📥 Format Dashboard Titles request")
        print(f"   filepath: {filepath}")
        print(f"   fontname: {fontname}")
        print(f"   fontsize: {fontsize}")
        print(f"   fontcolor: {fontcolor}")
        print(f"   bold: {bold}")
        
        if not filepath or not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        # Create a unique job ID
        job_id = str(uuid.uuid4())[:8]
        jobs[job_id] = {'status': 'processing', 'progress': 0, 'message': 'Formatting dashboard titles...'}
        
        # Process in background
        def process():
            try:
                import zipfile
                import xml.etree.ElementTree as ET
                
                # Create temp directory
                temp_dir = f"/tmp/title_formatter_{job_id}"
                if os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir)
                os.makedirs(temp_dir)
                
                # Extract workbook
                with zipfile.ZipFile(filepath, 'r') as zip_ref:
                    zip_ref.extractall(temp_dir)
                
                # Find .twb file
                twb_file = None
                for file in os.listdir(temp_dir):
                    if file.endswith('.twb'):
                        twb_file = os.path.join(temp_dir, file)
                        break
                
                if not twb_file:
                    jobs[job_id]['status'] = 'error'
                    jobs[job_id]['message'] = 'No .twb file found'
                    return
                
                # Parse and modify XML
                tree = ET.parse(twb_file)
                root = tree.getroot()
                
                dashboards_formatted = 0
                
                # Process each dashboard
                for dashboard in root.findall('.//dashboard'):
                    # Format layout-options title
                    layout_opts = dashboard.find('./layout-options')
                    if layout_opts is not None:
                        title_elem = layout_opts.find('./title')
                        if title_elem is not None:
                            fmt_text = title_elem.find('.//formatted-text')
                            if fmt_text is not None:
                                runs = fmt_text.findall('.//run')
                                for run in runs:
                                    run.set('fontname', fontname)
                                    run.set('fontsize', fontsize)
                                    run.set('fontcolor', fontcolor)
                                    if bold:
                                        run.set('bold', 'true')
                                dashboards_formatted += 1
                    
                    # Format zone-based titles (Calibre Medium 30)
                    for formatted_text in dashboard.findall('.//formatted-text'):
                        runs = formatted_text.findall('.//run')
                        if runs:
                            first_run = runs[0]
                            fn = first_run.get('fontname', 'default')
                            fs = first_run.get('fontsize', 'default')
                            
                            # Match primary target or similar titles
                            if (fn == 'Calibre Medium' and fs == '30') or \
                               (fn == 'Calibre' and fs in ['15', '24', '30']):
                                for run in runs:
                                    run.set('fontname', fontname)
                                    run.set('fontsize', fontsize)
                                    run.set('fontcolor', fontcolor)
                                    if bold:
                                        run.set('bold', 'true')
                
                # Save modified XML
                tree.write(twb_file, encoding='utf-8', xml_declaration=True)
                
                # Generate output path
                formatted_workbooks_path = os.path.expanduser("~/Desktop/Formatted_Workbooks")
                os.makedirs(formatted_workbooks_path, exist_ok=True)
                
                if original_filename:
                    output_filename = original_filename.rsplit('.', 1)[0] + f'_TITLES_{fontname.replace(" ", "_")}_{fontsize}pt_{fontcolor.replace("#", "")}.twbx'
                else:
                    output_filename = f'formatted_{job_id}_TITLES_{fontname.replace(" ", "_")}_{fontsize}pt_{fontcolor.replace("#", "")}.twbx'
                
                output_path = os.path.join(formatted_workbooks_path, output_filename)
                
                # Repackage as .twbx
                with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
                    for root_dir, dirs, files in os.walk(temp_dir):
                        for file in files:
                            file_path = os.path.join(root_dir, file)
                            arcname = os.path.relpath(file_path, temp_dir)
                            zip_ref.write(file_path, arcname)
                
                # Cleanup
                shutil.rmtree(temp_dir)
                
                jobs[job_id] = {
                    'status': 'completed',
                    'message': f'Successfully formatted {dashboards_formatted} dashboards',
                    'output_file': output_path,
                    'output_filename': output_filename,
                    'dashboards_formatted': dashboards_formatted
                }
                print(f"✅ Job {job_id} completed - {dashboards_formatted} dashboards formatted")
                
            except Exception as e:
                print(f"❌ Error in format_dashboard_titles: {e}")
                traceback.print_exc()
                jobs[job_id] = {
                    'status': 'error',
                    'message': str(e),
                    'traceback': traceback.format_exc()
                }
        
        thread = threading.Thread(target=process)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'success': True,
            'job_id': job_id
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    # Suppress all Flask startup messages
    import io
    import contextlib
    
    print("🚀 Starting Flask app on http://127.0.0.1:5555")
    
    # Silence Flask's startup logs
    f = io.StringIO()
    with contextlib.redirect_stdout(f):
        try:
            app.run(debug=False, host='127.0.0.1', port=5555, use_reloader=False)
        except Exception as e:
            print(f"❌ Error starting Flask: {e}")
            traceback.print_exc()
